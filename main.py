"""
Single-file Face Recognition Attendance System (Python 3.10)
- Camera via OpenCV
- Face recognition via face_recognition
- Data stored in SQLite
- Minimal Tkinter UI for adding students and starting attendance sessions

Run: py -3.10 main.py
"""
import csv
import json
import logging
import os
import shutil
import sqlite3
import sys
import threading
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import cv2
import face_recognition
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

DB_PATH = Path("attendance.db")
MAX_FACES_PER_FRAME = 20  # Can recognize up to 20 faces per frame (10+ students + teacher + buffer)
ENCODING_TOLERANCE = 0.5  # lower is stricter
CAMERA_INDEX = 1
ATTENDANCE_FRAME_SCALE = 0.5
ATTENDANCE_PROCESS_EVERY_N_FRAMES = 2
ATTENDANCE_CAMERA_WIDTH = 640
ATTENDANCE_CAMERA_HEIGHT = 480
UNKNOWN_VISITOR_SAVE_COOLDOWN_SEC = 2.0
ATTENDANCE_LOG_PATH = Path("attendance_debug.log")

# Face detector settings
# - "cnn" is intended to run on NVIDIA GPU *if* dlib was compiled with CUDA.
# - If CUDA isn't available, the code will automatically fall back to HOG.
ATTENDANCE_FACE_DETECTION_MODEL = os.getenv("ATTENDANCE_FACE_DETECTION_MODEL", "cnn").strip().lower()
ATTENDANCE_FACE_DETECTION_UPSAMPLE = int(os.getenv("ATTENDANCE_FACE_DETECTION_UPSAMPLE", "0"))
ATTENDANCE_CNN_FRAME_SCALE = float(os.getenv("ATTENDANCE_CNN_FRAME_SCALE", "0.25"))

LOGGER = logging.getLogger("attendance_app")
if not LOGGER.handlers:
    LOGGER.setLevel(logging.INFO)
    _log_handler = logging.FileHandler(ATTENDANCE_LOG_PATH, encoding="utf-8")
    _log_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] [%(threadName)s] %(message)s"))
    LOGGER.addHandler(_log_handler)


class Database:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self._ensure_db()

    def _ensure_db(self) -> None:
        cur = self.conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                encodings TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS teachers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                encodings TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS class_students (
                class_id INTEGER NOT NULL,
                student_id INTEGER NOT NULL,
                PRIMARY KEY (class_id, student_id),
                FOREIGN KEY (class_id) REFERENCES classes(id),
                FOREIGN KEY (student_id) REFERENCES students(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                class_id INTEGER NOT NULL,
                timestamp TEXT NOT NULL,
                FOREIGN KEY (student_id) REFERENCES students(id),
                FOREIGN KEY (class_id) REFERENCES classes(id)
            )
            """
        )
        # Migration: add session_number if the column is missing (for per-session history)
        try:
            cur.execute("SELECT session_number FROM attendance LIMIT 1")
        except sqlite3.OperationalError:
            cur.execute("ALTER TABLE attendance ADD COLUMN session_number INTEGER")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS teacher_attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_id INTEGER NOT NULL,
                class_id INTEGER NOT NULL,
                timestamp TEXT NOT NULL,
                session_number INTEGER,
                FOREIGN KEY (teacher_id) REFERENCES teachers(id),
                FOREIGN KEY (class_id) REFERENCES classes(id)
            )
            """
        )
        # Migration: add session_number if missing
        try:
            cur.execute("SELECT session_number FROM teacher_attendance LIMIT 1")
        except sqlite3.OperationalError:
            cur.execute("ALTER TABLE teacher_attendance ADD COLUMN session_number INTEGER")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS teacher_classes (
                teacher_id INTEGER NOT NULL,
                class_id INTEGER NOT NULL,
                PRIMARY KEY (teacher_id, class_id),
                FOREIGN KEY (teacher_id) REFERENCES teachers(id),
                FOREIGN KEY (class_id) REFERENCES classes(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS contact_info (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL UNIQUE,
                phone TEXT,
                email TEXT,
                address TEXT,
                guardian_name TEXT,
                guardian_phone TEXT,
                notes TEXT,
                FOREIGN KEY (student_id) REFERENCES students(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS unknown_visitors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL,
                session_number INTEGER,
                timestamp TEXT NOT NULL,
                image_path TEXT NOT NULL,
                encodings TEXT,
                visit_count INTEGER DEFAULT 1,
                last_seen TEXT,
                FOREIGN KEY (class_id) REFERENCES classes(id)
            )
            """
        )
        # Migration: add session_number if missing
        try:
            cur.execute("SELECT session_number FROM unknown_visitors LIMIT 1")
        except sqlite3.OperationalError:
            cur.execute("ALTER TABLE unknown_visitors ADD COLUMN session_number INTEGER")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS class_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL,
                session_number INTEGER NOT NULL,
                session_date TEXT NOT NULL,
                start_time TEXT,
                end_time TEXT,
                FOREIGN KEY (class_id) REFERENCES classes(id)
            )
            """
        )
        self.conn.commit()

    def add_student(self, name: str, encodings: List[np.ndarray]) -> int:
        if not encodings:
            raise ValueError("No face encodings found for this student.")
        enc_json = json.dumps([enc.tolist() for enc in encodings])
        cur = self.conn.cursor()
        cur.execute("INSERT INTO students (name, encodings) VALUES (?, ?)", (name, enc_json))
        student_id = cur.lastrowid
        self.conn.commit()
        return student_id

    def add_teacher(self, name: str, encodings: List[np.ndarray]) -> int:
        if not encodings:
            raise ValueError("No face encodings found for this teacher.")
        enc_json = json.dumps([enc.tolist() for enc in encodings])
        cur = self.conn.cursor()
        cur.execute("INSERT INTO teachers (name, encodings) VALUES (?, ?)", (name, enc_json))
        teacher_id = cur.lastrowid
        self.conn.commit()
        return teacher_id

    def list_teachers(self) -> List[Tuple[int, str]]:
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM teachers ORDER BY id")
        return cur.fetchall()

    def get_teacher_encodings(self, teacher_id: int) -> List[np.ndarray]:
        cur = self.conn.cursor()
        cur.execute("SELECT encodings FROM teachers WHERE id=?", (teacher_id,))
        row = cur.fetchone()
        if not row:
            return []
        data = json.loads(row[0])
        return [np.array(enc) for enc in data]

    def get_all_teacher_encodings(self) -> Tuple[List[np.ndarray], List[int]]:
        teachers = self.list_teachers()
        encodings = []
        teacher_ids = []
        for tid, _ in teachers:
            encs = self.get_teacher_encodings(tid)
            encodings.extend(encs)
            teacher_ids.extend([tid] * len(encs))
        return encodings, teacher_ids

    def mark_teacher_attendance(
        self,
        class_id: int,
        teacher_id: int,
        ts: datetime,
        session_number: Optional[int] = None,
    ) -> None:
        ts_str = ts.isoformat(timespec="seconds")
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO teacher_attendance (teacher_id, class_id, timestamp, session_number) VALUES (?, ?, ?, ?)",
            (teacher_id, class_id, ts_str, session_number),
        )
        self.conn.commit()

    def save_unknown_visitor(
        self,
        class_id: int,
        image_path: str,
        ts: datetime,
        encoding: Optional[np.ndarray] = None,
        session_number: Optional[int] = None,
    ) -> int:
        """Save or update unknown visitor. Returns visitor ID."""
        ts_str = ts.isoformat(timespec="seconds")
        cur = self.conn.cursor()
        
        # Check if this face has been seen before in this class
        if encoding is not None:
            existing_visitors = cur.execute(
                "SELECT id, encodings, visit_count FROM unknown_visitors WHERE class_id=?",
                (class_id,),
            ).fetchall()
            
            # Check against each existing visitor
            for visitor_id, enc_json, visit_count in existing_visitors:
                if enc_json:
                    stored_encs = [np.array(e) for e in json.loads(enc_json)]
                    distances = face_recognition.face_distance(stored_encs, encoding)
                    if len(distances) > 0 and min(distances) <= 0.5:
                        # This is a repeat visitor!
                        cur.execute(
                            "UPDATE unknown_visitors SET visit_count=?, last_seen=? WHERE id=?",
                            (visit_count + 1, ts_str, visitor_id)
                        )
                        self.conn.commit()
                        return visitor_id
        
        # New visitor - save with encoding
        enc_json = json.dumps([encoding.tolist()]) if encoding is not None else None
        cur.execute(
            """INSERT INTO unknown_visitors 
               (class_id, session_number, timestamp, image_path, encodings, visit_count, last_seen) 
               VALUES (?, ?, ?, ?, ?, 1, ?)""",
            (class_id, session_number, ts_str, image_path, enc_json, ts_str),
        )
        visitor_id = cur.lastrowid
        self.conn.commit()
        return visitor_id

    def delete_class_session(self, class_id: int, session_number: int) -> None:
        """Delete a class session (used when teacher never arrives and class is cancelled)."""
        cur = self.conn.cursor()
        cur.execute(
            "DELETE FROM class_sessions WHERE class_id=? AND session_number=?",
            (class_id, session_number),
        )
        # If any attendance rows were recorded for this session (rare), remove them too.
        cur.execute(
            "DELETE FROM attendance WHERE class_id=? AND session_number=?",
            (class_id, session_number),
        )
        # Teacher/visitor tables might not have session numbers historically; keep safe.
        try:
            cur.execute(
                "DELETE FROM teacher_attendance WHERE class_id=? AND session_number=?",
                (class_id, session_number),
            )
        except Exception:
            pass
        try:
            cur.execute(
                "DELETE FROM unknown_visitors WHERE class_id=? AND session_number=?",
                (class_id, session_number),
            )
        except Exception:
            pass
        self.conn.commit()

    def get_class_total_sessions(self, class_id: int) -> int:
        cur = self.conn.cursor()
        row = cur.execute(
            "SELECT COUNT(*) FROM class_sessions WHERE class_id=?",
            (class_id,),
        ).fetchone()
        return int(row[0] or 0)

    @staticmethod
    def warning_from_absences(absent_lectures: int) -> str:
        if absent_lectures >= 4:
            return "Last warning"
        if absent_lectures == 3:
            return "2nd warning"
        if absent_lectures == 2:
            return "1st warning"
        return ""

    def get_warning_summary_for_class(self, class_id: int) -> List[Tuple[int, str, int, int, int, str]]:
        """Per-class warning summary.

        Returns rows: (student_id, student_name, total_lectures, attended_lectures, absent_lectures, warning)
        """
        roster = self.list_class_students(class_id)
        total_lectures = self.get_class_total_sessions(class_id)

        cur = self.conn.cursor()
        rows = cur.execute(
            "SELECT student_id, session_number, timestamp FROM attendance WHERE class_id=?",
            (class_id,),
        ).fetchall()

        attended_keys_by_student: Dict[int, set] = defaultdict(set)
        all_keys: set = set()

        for student_id, session_number, ts_str in rows:
            if session_number is not None and int(session_number) > 0:
                key = f"s{int(session_number)}"
            else:
                # Fallback for older data without session_number: treat each distinct date as a lecture.
                date_part = str(ts_str).split("T")[0] if ts_str else ""
                key = f"d{date_part}" if date_part else "dunknown"
            attended_keys_by_student[int(student_id)].add(key)
            all_keys.add(key)

        if total_lectures <= 0:
            total_lectures = len(all_keys)

        summary: List[Tuple[int, str, int, int, int, str]] = []
        for student_id, student_name in roster:
            attended = len(attended_keys_by_student.get(int(student_id), set()))
            absent = max(0, total_lectures - attended)
            warning = self.warning_from_absences(absent)
            summary.append((int(student_id), str(student_name), int(total_lectures), int(attended), int(absent), warning))
        return summary

    def get_all_warnings_summary(self) -> List[Tuple[int, str, str, int, int, int, str]]:
        """Get warnings summary for ALL classes.

        Returns rows: (student_id, student_name, class_name, total_lectures, attended, absent, warning)
        """
        all_classes = self.list_classes()
        summary = []
        for class_id, class_name in all_classes:
            class_summary = self.get_warning_summary_for_class(class_id)
            for student_id, student_name, total_lect, attended, absent, warning in class_summary:
                summary.append((student_id, student_name, class_name, total_lect, attended, absent, warning))
        return summary

    
    def create_class_session(self, class_id: int, session_date: str, 
                            start_time: str = "") -> int:
        """Create a new session for a class and return session_number"""
        cur = self.conn.cursor()
        
        # Get the next session number for this class
        max_session = cur.execute(
            "SELECT MAX(session_number) FROM class_sessions WHERE class_id=?",
            (class_id,)
        ).fetchone()[0]
        
        next_session = (max_session or 0) + 1
        
        cur.execute(
            """INSERT INTO class_sessions 
               (class_id, session_number, session_date, start_time) 
               VALUES (?, ?, ?, ?)""",
            (class_id, next_session, session_date, start_time)
        )
        session_id = cur.lastrowid
        self.conn.commit()
        return next_session

    def get_teacher_classes(self, teacher_id: int) -> List[int]:
        cur = self.conn.cursor()
        cur.execute("SELECT class_id FROM teacher_classes WHERE teacher_id=?", (teacher_id,))
        return [row[0] for row in cur.fetchall()]

    def update_teacher_classes(self, teacher_id: int, class_ids: List[int]) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM teacher_classes WHERE teacher_id=?", (teacher_id,))
        if class_ids:
            cur.executemany(
                "INSERT INTO teacher_classes (teacher_id, class_id) VALUES (?, ?)",
                [(teacher_id, cid) for cid in class_ids],
            )
        self.conn.commit()

    def delete_teacher(self, teacher_id: int) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM teacher_classes WHERE teacher_id=?", (teacher_id,))
        cur.execute("DELETE FROM teachers WHERE id=?", (teacher_id,))
        cur.execute("DELETE FROM teacher_attendance WHERE teacher_id=?", (teacher_id,))
        self.conn.commit()

    def list_students(self) -> List[Tuple[int, str]]:
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM students ORDER BY id")
        return cur.fetchall()

    def update_student(self, student_id: int, name: str, class_ids: List[int]) -> None:
        cur = self.conn.cursor()
        cur.execute("UPDATE students SET name=? WHERE id=?", (name, student_id))
        cur.execute("DELETE FROM class_students WHERE student_id=?", (student_id,))
        if class_ids:
            cur.executemany(
                "INSERT INTO class_students (class_id, student_id) VALUES (?, ?)",
                [(cid, student_id) for cid in class_ids],
            )
        self.conn.commit()

    def delete_student(self, student_id: int) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM class_students WHERE student_id=?", (student_id,))
        cur.execute("DELETE FROM students WHERE id=?", (student_id,))
        cur.execute("DELETE FROM attendance WHERE student_id=?", (student_id,))
        self.conn.commit()

    def get_student_classes(self, student_id: int) -> List[int]:
        cur = self.conn.cursor()
        cur.execute("SELECT class_id FROM class_students WHERE student_id=?", (student_id,))
        return [row[0] for row in cur.fetchall()]

    def get_student_encodings(self, student_id: int) -> List[np.ndarray]:
        cur = self.conn.cursor()
        cur.execute("SELECT encodings FROM students WHERE id=?", (student_id,))
        row = cur.fetchone()
        if not row:
            return []
        data = json.loads(row[0])
        return [np.array(enc) for enc in data]

    def create_class_definition(self, name: str, student_ids: List[int]) -> int:
        cur = self.conn.cursor()
        cur.execute("INSERT INTO classes (name, start_time, created_at) VALUES (?, ?, ?)", (name, "", datetime.now().isoformat(timespec="seconds")))
        class_id = cur.lastrowid
        if student_ids:
            cur.executemany(
                "INSERT OR IGNORE INTO class_students (class_id, student_id) VALUES (?, ?)",
                [(class_id, sid) for sid in student_ids],
            )
        self.conn.commit()
        return class_id

    def list_classes(self) -> List[Tuple[int, str]]:
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM classes ORDER BY name")
        return cur.fetchall()

    def delete_class(self, class_id: int) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM class_students WHERE class_id=?", (class_id,))
        cur.execute("DELETE FROM classes WHERE id=?", (class_id,))
        cur.execute("DELETE FROM attendance WHERE class_id=?", (class_id,))
        self.conn.commit()

    def update_class_students(self, class_id: int, student_ids: List[int]) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM class_students WHERE class_id=?", (class_id,))
        if student_ids:
            cur.executemany(
                "INSERT INTO class_students (class_id, student_id) VALUES (?, ?)",
                [(class_id, sid) for sid in student_ids],
            )
        self.conn.commit()

    def list_class_students(self, class_id: int) -> List[Tuple[int, str]]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT s.id, s.name
            FROM students s
            JOIN class_students cs ON cs.student_id = s.id
            WHERE cs.class_id = ?
            ORDER BY s.name
            """,
            (class_id,),
        )
        return cur.fetchall()

    def mark_attendance(self, class_id: int, student_id: int, ts: datetime, 
                       session_number: Optional[int] = None) -> None:
        ts_str = ts.isoformat(timespec="seconds")
        cur = self.conn.cursor()
        
        # Add session_number column if it doesn't exist
        try:
            cur.execute("SELECT session_number FROM attendance LIMIT 1")
        except:
            cur.execute("ALTER TABLE attendance ADD COLUMN session_number INTEGER")
        
        cur.execute(
            """INSERT INTO attendance 
               (student_id, class_id, timestamp, session_number) 
               VALUES (?, ?, ?, ?)""",
            (student_id, class_id, ts_str, session_number),
        )
        self.conn.commit()

    def attendance_for_class(self, class_id: int) -> List[Tuple[str, str]]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT s.name, a.timestamp
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            WHERE a.class_id = ?
            ORDER BY a.timestamp
            """,
            (class_id,),
        )
        return cur.fetchall()

    def absent_for_class(self, class_id: int) -> List[str]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT s.name
            FROM students s
            JOIN class_students cs ON cs.student_id = s.id
            WHERE cs.class_id = ? AND s.id NOT IN (
                SELECT student_id FROM attendance WHERE class_id = ?
            )
            ORDER BY s.name
            """,
            (class_id, class_id),
        )
        rows = cur.fetchall()
        return [r[0] for r in rows]
    
    def get_all_attendance_dates(self) -> List[str]:
        """Get all unique dates that have attendance records"""
        cur = self.conn.cursor()
        cur.execute("""
            SELECT DISTINCT date(timestamp) as attendance_date
            FROM attendance
            ORDER BY attendance_date DESC
        """)
        return [row[0] for row in cur.fetchall()]
    
    def get_classes_for_date(self, date_str: str) -> List[Tuple[int, str]]:
        """Get all classes that had attendance on a specific date"""
        date_start = f"{date_str}T00:00:00"
        date_end = f"{date_str}T23:59:59"
        cur = self.conn.cursor()
        cur.execute("""
            SELECT DISTINCT c.id, c.name
            FROM classes c
            JOIN attendance a ON c.id = a.class_id
            WHERE a.timestamp >= ? AND a.timestamp <= ?
            ORDER BY c.name
        """, (date_start, date_end))
        return cur.fetchall()


class FaceAttendanceApp:
    def __init__(self) -> None:
        self.db = Database(DB_PATH)
        self.root = tk.Tk()
        self.root.title("✨ TruePresence - Smart Attendance System")
        self.root.geometry("1400x850")
        self.root.resizable(True, True)
        self.roster_index: Dict[int, int] = {}
        
        # Modern color palette
        self.colors = {
            'primary': '#667eea',        # Purple-blue gradient start
            'primary_dark': '#764ba2',   # Purple-blue gradient end
            'secondary': '#f093fb',      # Pink gradient
            'accent': '#4facfe',         # Bright blue
            'success': '#11d17e',        # Vibrant green
            'warning': '#ffa502',        # Orange
            'danger': '#ff4757',         # Red
            'info': '#00d2ff',           # Cyan
            'dark': '#2c3e50',           # Dark blue-gray
            'light': '#f8f9fa',          # Light gray
            'bg_main': '#ffffff',        # White background
            'bg_secondary': '#f7f8fc',   # Light purple-gray
            'bg_card': '#ffffff',        # Card background
            'text_primary': '#2c3e50',   # Dark text
            'text_secondary': '#7f8c8d', # Light text
            'border': '#e1e8ed',         # Border color
            'shadow': '#0000001a'        # Shadow color
        }
        
        self._init_style()
        self._build_ui()
        self._apply_theme()

    def _init_style(self) -> None:
        """Set modern fonts and styling."""
        try:
            # Modern font stack
            self.root.option_add("*Font", "{Segoe UI} 10")
            self.root.option_add("*Button.Font", "{Segoe UI} 10 bold")
            self.root.option_add("*Label.Font", "{Segoe UI} 10")
            self.root.option_add("*Entry.Font", "{Segoe UI} 10")
            self.root.option_add("*Listbox.Font", "{Segoe UI} 10")
            self.root.option_add("*Menu.Font", "{Segoe UI} 10")
            
            # Set main window background
            self.root.configure(bg=self.colors['bg_secondary'])
        except Exception:
            pass
    
    def _create_gradient_frame(self, parent, height=100, color1='#667eea', color2='#764ba2'):
        """Create a frame with simulated gradient using multiple frames."""
        gradient_frame = tk.Frame(parent, height=height, bg=color1)
        return gradient_frame
    
    def _create_styled_button(self, parent, text, command, bg_color, fg_color='white', width=None):
        """Create a modern styled button with hover effect."""
        btn = tk.Button(parent, text=text, command=command, 
                       bg=bg_color, fg=fg_color, 
                       font=("Segoe UI", 10, "bold"),
                       bd=0, relief='flat',
                       cursor='hand2',
                       activebackground=self._darken_color(bg_color),
                       activeforeground=fg_color)
        if width:
            btn.config(width=width)
        
        # Hover effect
        def on_enter(e):
            btn['bg'] = self._darken_color(bg_color)
        
        def on_leave(e):
            btn['bg'] = bg_color
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn
    
    def _darken_color(self, color, factor=0.85):
        """Darken a hex color."""
        try:
            color = color.lstrip('#')
            rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            darkened = tuple(int(c * factor) for c in rgb)
            return '#{:02x}{:02x}{:02x}'.format(*darkened)
        except:
            return color
    
    def _create_stat_card(self, parent, title, value, color, stat_type):
        """Create a beautiful stat card with icon and value."""
        card = tk.Frame(parent, bg='white', relief='flat', bd=0)
        card.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 15))
        
        # Color accent bar on left
        accent = tk.Frame(card, bg=color, width=6)
        accent.pack(side=tk.LEFT, fill="y")
        
        # Card content
        content = tk.Frame(card, bg='white')
        content.pack(side=tk.LEFT, fill="both", expand=True, padx=20, pady=20)
        
        # Title
        tk.Label(content, text=title, 
                font=("Segoe UI", 11, "bold"), 
                bg='white', 
                fg=self.colors['text_secondary']).pack(anchor='w')
        
        # Value (large number)
        value_label = tk.Label(content, text=value, 
                              font=("Segoe UI", 32, "bold"), 
                              bg='white', 
                              fg=color)
        value_label.pack(anchor='w', pady=(5, 0))
        
        # Store reference for updating
        if stat_type == 'present':
            self.stat_present_label = value_label
        elif stat_type == 'late':
            self.stat_late_label = value_label
        elif stat_type == 'absent':
            self.stat_absent_label = value_label
        elif stat_type == 'teachers':
            self.stat_teachers_label = value_label
        elif stat_type == 'visitors':
            self.stat_visitors_label = value_label
        elif stat_type == 'classes':
            self.stat_classes_label = value_label
        
        return card

    def _build_ui(self) -> None:
        from tkinter import ttk
        pad = {"padx": 10, "pady": 8}
        
        # ============ MODERN HEADER ============
        header = tk.Frame(self.root, bg=self.colors['primary'], height=80)
        header.pack(side=tk.TOP, fill="x")
        header.pack_propagate(False)
        
        # Header content with icon and title
        header_content = tk.Frame(header, bg=self.colors['primary'])
        header_content.pack(expand=True)
        
        # App icon (emoji-style)
        icon_label = tk.Label(header_content, text="🎓", font=("Segoe UI Emoji", 40), bg=self.colors['primary'], fg='white')
        icon_label.pack(side=tk.LEFT, padx=15)
        
        # Title and subtitle
        title_frame = tk.Frame(header_content, bg=self.colors['primary'])
        title_frame.pack(side=tk.LEFT, padx=10)
        
        title_label = tk.Label(title_frame, text="TruePresence", 
                              font=("Segoe UI", 24, "bold"), 
                              bg=self.colors['primary'], 
                              fg='white')
        title_label.pack(anchor='w')
        
        subtitle_label = tk.Label(title_frame, text="Smart Face Recognition Attendance System", 
                     font=("Segoe UI", 11), 
                     bg=self.colors['primary'], 
                     fg='#f5f5f5')
        subtitle_label.pack(anchor='w')
        
        # Main container
        main_container = tk.Frame(self.root, bg=self.colors['bg_secondary'])
        main_container.pack(side=tk.TOP, fill="both", expand=True)
        
        # ============ LEFT PANEL - STUDENT MANAGEMENT ============
        left_panel = tk.Frame(main_container, bg=self.colors['bg_secondary'])
        left_panel.pack(side=tk.LEFT, fill="y", expand=False, padx=15, pady=15)
        left_panel.pack_propagate(False)
        left_panel.config(width=380)
        
        # Add Student Section - Modern Card Style
        add_frame = tk.Frame(left_panel, bg=self.colors['bg_card'], relief='flat', bd=0)
        add_frame.pack(fill="x", pady=(0, 12))
        
        # Card header with gradient feel
        card_header = tk.Frame(add_frame, bg=self.colors['primary'], height=45)
        card_header.pack(fill="x")
        card_header.pack_propagate(False)
        
        tk.Label(card_header, text="➕ Add New Student", 
                font=("Segoe UI", 12, "bold"), 
                bg=self.colors['primary'], 
                fg='white').pack(side=tk.LEFT, padx=15, pady=10)
        
        # Card body
        card_body = tk.Frame(add_frame, bg=self.colors['bg_card'])
        card_body.pack(fill="x", padx=15, pady=15)
        
        tk.Label(card_body, text="👤 Student Name", 
                font=("Segoe UI", 10, "bold"), 
                bg=self.colors['bg_card'], 
                fg=self.colors['text_primary']).pack(**pad, anchor="w")
        
        self.student_name_var = tk.StringVar()
        name_entry = tk.Entry(card_body, textvariable=self.student_name_var, 
                             font=("Segoe UI", 11), bd=1, relief='solid')
        name_entry.pack(fill="x", **pad, ipady=8)
        
        # Styled buttons with icons
        self._create_styled_button(card_body, "📁 Choose Images", self._select_images, 
                                   self.colors['info'], 'white').pack(**pad, fill="x", ipady=8)
        self._create_styled_button(card_body, "📷 Use Camera", self._capture_from_camera, 
                                   self.colors['accent'], 'white').pack(**pad, fill="x", ipady=8)
        self._create_styled_button(card_body, "💾 Save Student", self._save_student, 
                                   self.colors['success'], 'white').pack(**pad, fill="x", ipady=8)
        
        self.selected_images: List[str] = []
        self.captured_encodings: List[np.ndarray] = []
        self.adding_student_id: Optional[int] = None
        
        # Manage Students Section - Modern Card Style
        student_mgmt_frame = tk.Frame(left_panel, bg=self.colors['bg_card'], relief='flat', bd=0)
        student_mgmt_frame.pack(fill="both", expand=True, pady=(0, 0))
        
        # Card header
        mgmt_header = tk.Frame(student_mgmt_frame, bg=self.colors['primary_dark'], height=45)
        mgmt_header.pack(fill="x")
        mgmt_header.pack_propagate(False)
        
        tk.Label(mgmt_header, text="👥 Manage Students", 
                font=("Segoe UI", 12, "bold"), 
                bg=self.colors['primary_dark'], 
                fg='white').pack(side=tk.LEFT, padx=15, pady=10)
        
        # Card body
        mgmt_body = tk.Frame(student_mgmt_frame, bg=self.colors['bg_card'])
        mgmt_body.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Modern search box
        search_frame = tk.Frame(mgmt_body, bg=self.colors['bg_card'])
        search_frame.pack(fill="x", pady=(0, 12))
        
        tk.Label(search_frame, text="🔍", font=("Segoe UI Emoji", 14), 
                bg=self.colors['bg_card']).pack(side=tk.LEFT, padx=(0, 5))
        
        self.student_search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.student_search_var, 
                               font=("Segoe UI", 10), bd=1, relief='solid')
        search_entry.pack(side=tk.LEFT, fill="x", expand=True, ipady=6)
        search_entry.insert(0, "Search students...")
        search_entry.config(fg=self.colors['text_secondary'])
        
        # Bind focus events for placeholder behavior
        def on_focus_in(e):
            if search_entry.get() == "Search students...":
                search_entry.delete(0, tk.END)
                search_entry.config(fg=self.colors['text_primary'])
        
        def on_focus_out(e):
            if not search_entry.get():
                search_entry.insert(0, "Search students...")
                search_entry.config(fg=self.colors['text_secondary'])
        
        search_entry.bind("<FocusIn>", on_focus_in)
        search_entry.bind("<FocusOut>", on_focus_out)
        # Defer trace until listbox exists to avoid attribute error
        def _attach_search_trace():
            if hasattr(self, 'student_listbox'):
                self.student_search_var.trace("w", lambda *args: self._search_students())
        self.root.after(100, _attach_search_trace)
        
        tk.Label(mgmt_body, text="Students List", 
                font=("Segoe UI", 9, "bold"), 
                bg=self.colors['bg_card'], 
                fg=self.colors['text_secondary']).pack(anchor="w", pady=(8, 5))
        
        student_list_frame = tk.Frame(mgmt_body, bg=self.colors['bg_card'])
        student_list_frame.pack(fill="both", expand=True)
        
        student_scrollbar = ttk.Scrollbar(student_list_frame)
        student_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.student_listbox = tk.Listbox(student_list_frame, height=10, 
                                         yscrollcommand=student_scrollbar.set, 
                                         font=("Segoe UI", 10), 
                                         bd=1, relief='solid',
                                         selectbackground=self.colors['primary'],
                                         selectforeground='white',
                                         bg='white')
        student_scrollbar.config(command=self.student_listbox.yview)
        self.student_listbox.pack(fill="both", expand=True, side=tk.LEFT)
        self.student_listbox.bind("<<ListboxSelect>>", self._on_student_selected)
        
        # Info display with modern look
        tk.Label(mgmt_body, text="ℹ️ Information", 
                font=("Segoe UI", 9, "bold"), 
                bg=self.colors['bg_card'], 
                fg=self.colors['text_secondary']).pack(anchor="w", pady=(12, 5))
        
        info_container = tk.Frame(mgmt_body, bg=self.colors['bg_secondary'], 
                                 bd=1, relief='solid')
        info_container.pack(fill="x", pady=(0, 12))
        
        self.student_info_var = tk.StringVar(value="Select a student to view details")
        tk.Label(info_container, textvariable=self.student_info_var, 
                font=("Segoe UI", 9), wraplength=320, justify="left", 
                bg=self.colors['bg_secondary'], 
                fg=self.colors['text_primary']).pack(padx=12, pady=12)
        
        # Action buttons with modern styling
        btn_frame = tk.Frame(mgmt_body, bg=self.colors['bg_card'])
        btn_frame.pack(fill="x")
        
        self._create_styled_button(btn_frame, "✏️ Edit", self._edit_selected_student, 
                                   self.colors['info'], 'white', width=9).pack(side=tk.LEFT, padx=(0, 5), ipady=6)
        self._create_styled_button(btn_frame, "🗑️ Delete", self._delete_selected_student, 
                                   self.colors['danger'], 'white', width=9).pack(side=tk.LEFT, padx=5, ipady=6)
        self._create_styled_button(btn_frame, "🔄 Refresh", self._refresh_student_list, 
                                   self.colors['dark'], 'white', width=9).pack(side=tk.LEFT, padx=5, ipady=6)
        
        # ============ RIGHT PANEL - CLASSES, ATTENDANCE, ROSTER, LOGS ============
        right_panel = tk.Frame(main_container, bg=self.colors['bg_secondary'])
        right_panel.pack(side=tk.RIGHT, fill="both", expand=True, padx=(0, 15), pady=15)
        
        # Create a modern notebook (tabs) with custom styling
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure tab style
        style.configure('Modern.TNotebook', background=self.colors['bg_secondary'], borderwidth=0)
        style.configure('Modern.TNotebook.Tab', 
                       background=self.colors['bg_card'],
                       foreground=self.colors['text_primary'],
                       padding=[20, 12],
                       font=('Segoe UI', 10, 'bold'),
                       borderwidth=0)
        style.map('Modern.TNotebook.Tab',
                 background=[('selected', self.colors['primary'])],
                 foreground=[('selected', 'white')],
                 expand=[('selected', [1, 1, 1, 0])])
        
        notebook = ttk.Notebook(right_panel, style='Modern.TNotebook')
        notebook.pack(fill="both", expand=True)
        
        # Tab 1: Dashboard with Beautiful Stats Cards
        dashboard_tab = tk.Frame(notebook, bg=self.colors['bg_secondary'])
        notebook.add(dashboard_tab, text="📊 Dashboard")
        
        # Welcome banner
        welcome_frame = tk.Frame(dashboard_tab, bg=self.colors['primary'], height=100)
        welcome_frame.pack(fill="x", padx=20, pady=(20, 15))
        welcome_frame.pack_propagate(False)
        
        welcome_content = tk.Frame(welcome_frame, bg=self.colors['primary'])
        welcome_content.pack(expand=True)
        
        tk.Label(welcome_content, text="👋 Welcome Back!", 
                font=("Segoe UI", 18, "bold"), 
                bg=self.colors['primary'], 
                fg='white').pack(anchor='w')
        
        tk.Label(welcome_content, text=datetime.now().strftime("Today is %A, %B %d, %Y"), 
            font=("Segoe UI", 11), 
            bg=self.colors['primary'], 
            fg='#f5f5f5').pack(anchor='w', pady=(5, 0))
        
        # Stats cards container
        stats_container = tk.Frame(dashboard_tab, bg=self.colors['bg_secondary'])
        stats_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Row 1: Main statistics
        stats_row1 = tk.Frame(stats_container, bg=self.colors['bg_secondary'])
        stats_row1.pack(fill="x", pady=(0, 15))
        
        # Create stat cards with icons and colors
        self._create_stat_card(stats_row1, "👥 Present", "0", self.colors['success'], 'present')
        self._create_stat_card(stats_row1, "⏰ Late", "0", self.colors['warning'], 'late')
        self._create_stat_card(stats_row1, "❌ Absent", "0", self.colors['danger'], 'absent')
        
        # Row 2: Additional statistics
        stats_row2 = tk.Frame(stats_container, bg=self.colors['bg_secondary'])
        stats_row2.pack(fill="x", pady=(0, 15))
        
        self._create_stat_card(stats_row2, "👨‍🏫 Teachers", "0", self.colors['info'], 'teachers')
        self._create_stat_card(stats_row2, "👤 Unknown", "0", self.colors['primary_dark'], 'visitors')
        self._create_stat_card(stats_row2, "📚 Classes", "0", self.colors['accent'], 'classes')
        
        # Action buttons
        action_frame = tk.Frame(stats_container, bg=self.colors['bg_secondary'])
        action_frame.pack(fill="x", pady=(10, 0))
        
        self._create_styled_button(action_frame, "🔄 Refresh Dashboard", self._refresh_dashboard, 
                                   self.colors['primary'], 'white').pack(side=tk.LEFT, padx=(0, 10), ipady=10, ipadx=20)
        
        self._create_styled_button(action_frame, "▶️ Start Attendance", 
                                   lambda: notebook.select(1), 
                                   self.colors['success'], 'white').pack(side=tk.LEFT, ipady=10, ipadx=20)
        
        # Tab 2: Classes & Start Attendance (Modern Design)
        classes_tab = tk.Frame(notebook, bg=self.colors['bg_secondary'])
        notebook.add(classes_tab, text="🎓 Classes")
        
        # Two-column layout
        left_col = tk.Frame(classes_tab, bg=self.colors['bg_secondary'])
        left_col.pack(side=tk.LEFT, fill="both", expand=True, padx=(20, 10), pady=20)
        
        right_col = tk.Frame(classes_tab, bg=self.colors['bg_secondary'])
        right_col.pack(side=tk.RIGHT, fill="both", expand=True, padx=(10, 20), pady=20)
        
        # Left: Manage Classes Card
        classes_card = tk.Frame(left_col, bg='white', relief='flat')
        classes_card.pack(fill="both", expand=True)
        
        classes_header = tk.Frame(classes_card, bg=self.colors['accent'], height=50)
        classes_header.pack(fill="x")
        classes_header.pack_propagate(False)
        
        tk.Label(classes_header, text="📚 Manage Classes", 
                font=("Segoe UI", 13, "bold"), 
                bg=self.colors['accent'], 
                fg='white').pack(side=tk.LEFT, padx=20, pady=12)
        
        classes_body = tk.Frame(classes_card, bg='white')
        classes_body.pack(fill="both", expand=True, padx=20, pady=20)
        
        class_list_frame = tk.Frame(classes_body, bg='white')
        class_list_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        class_scrollbar = ttk.Scrollbar(class_list_frame)
        class_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.class_listbox = tk.Listbox(class_list_frame, height=12, 
                                        yscrollcommand=class_scrollbar.set, 
                                        font=("Segoe UI", 11), 
                                        bd=1, relief='solid',
                                        selectbackground=self.colors['accent'],
                                        selectforeground='white',
                                        bg='white')
        class_scrollbar.config(command=self.class_listbox.yview)
        self.class_listbox.pack(fill="both", expand=True, side=tk.LEFT)
        self.class_listbox.bind("<<ListboxSelect>>", self._on_class_selected)
        
        # Store reference to notebook for switching tabs
        self.main_notebook = notebook
        
        class_btn_frame = tk.Frame(classes_body, bg='white')
        class_btn_frame.pack(fill="x")
        
        self._create_styled_button(class_btn_frame, "➕ New", self._create_class_dialog, 
                                   self.colors['success'], 'white', width=10).pack(side=tk.LEFT, padx=(0, 8), ipady=8)
        self._create_styled_button(class_btn_frame, "✏️ Edit", self._edit_class_dialog, 
                                   self.colors['info'], 'white', width=10).pack(side=tk.LEFT, padx=8, ipady=8)
        self._create_styled_button(class_btn_frame, "🗑️ Delete", self._delete_class_dialog, 
                                   self.colors['danger'], 'white', width=10).pack(side=tk.LEFT, padx=8, ipady=8)
        
        # Right: Start Attendance Card
        attend_card = tk.Frame(right_col, bg='white', relief='flat')
        attend_card.pack(fill="both", expand=True)
        
        attend_header = tk.Frame(attend_card, bg=self.colors['success'], height=50)
        attend_header.pack(fill="x")
        attend_header.pack_propagate(False)
        
        tk.Label(attend_header, text="▶️ Start Attendance", 
                font=("Segoe UI", 13, "bold"), 
                bg=self.colors['success'], 
                fg='white').pack(side=tk.LEFT, padx=20, pady=12)
        
        attend_body = tk.Frame(attend_card, bg='white')
        attend_body.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Form fields with modern styling
        def create_form_field(parent, label_text, var, entry_type='entry'):
            field_frame = tk.Frame(parent, bg='white')
            field_frame.pack(fill="x", pady=(0, 15))
            
            tk.Label(field_frame, text=label_text, 
                    font=("Segoe UI", 10, "bold"), 
                    bg='white', 
                    fg=self.colors['text_primary']).pack(anchor='w', pady=(0, 5))
            
            if entry_type == 'entry':
                entry = tk.Entry(field_frame, textvariable=var, 
                               font=("Segoe UI", 11), 
                               bd=1, relief='solid')
                entry.pack(fill="x", ipady=8)
                return entry
            elif entry_type == 'dropdown':
                dropdown = tk.OptionMenu(field_frame, var, "")
                dropdown.config(font=("Segoe UI", 10), bd=1, relief='solid', 
                              bg='white', activebackground=self.colors['bg_secondary'])
                dropdown.pack(fill="x", ipady=4)
                return dropdown
        
        self.class_select_var = tk.StringVar(value="")
        self.class_dropdown = create_form_field(attend_body, "📖 Select Class", 
                                                self.class_select_var, 'dropdown')
        
        refresh_btn = self._create_styled_button(attend_body, "🔄 Refresh Classes", 
                                                self._refresh_class_dropdown, 
                                                self.colors['info'], 'white')
        refresh_btn.pack(fill="x", pady=(0, 15), ipady=6)
        
        self.class_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        create_form_field(attend_body, "📅 Date", self.class_date_var)
        
        self.class_time_var = tk.StringVar(value="10:00 AM")
        create_form_field(attend_body, "🕐 Start Time", self.class_time_var)
        
        self.class_end_time_var = tk.StringVar(value="")
        create_form_field(attend_body, "🕐 End Time (optional)", self.class_end_time_var)
        
        # Camera settings
        camera_frame = tk.Frame(attend_body, bg='white')
        camera_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(camera_frame, text="📷 Camera Index", 
                font=("Segoe UI", 10, "bold"), 
                bg='white', 
                fg=self.colors['text_primary']).pack(anchor='w', pady=(0, 5))
        
        camera_input_frame = tk.Frame(camera_frame, bg='white')
        camera_input_frame.pack(fill="x")
        
        self.camera_index_var = tk.StringVar(value=str(CAMERA_INDEX))
        cam_entry = tk.Entry(camera_input_frame, textvariable=self.camera_index_var, 
                            width=8, font=("Segoe UI", 11), bd=1, relief='solid')
        cam_entry.pack(side=tk.LEFT, ipady=8, padx=(0, 10))
        
        self._create_styled_button(camera_input_frame, "🎥 Test", self._test_camera, 
                                   self.colors['warning'], 'white', width=12).pack(side=tk.LEFT, ipady=6)
        
        # Start button (prominent)
        start_btn = self._create_styled_button(attend_body, "▶️ START ATTENDANCE SESSION", 
                                              self._start_attendance_thread, 
                                              self.colors['success'], 'white')
        start_btn.config(font=("Segoe UI", 12, "bold"))
        start_btn.pack(fill="x", pady=(20, 0), ipady=15)
        
        # Tab 3: Roster & Logs (Redesigned)
        roster_tab = tk.Frame(notebook, bg=self.colors['bg_secondary'])
        notebook.add(roster_tab, text="📋 Roster & Logs")
        
        # Roster Section - Modern Card
        roster_card = tk.Frame(roster_tab, bg='white', relief='flat')
        roster_card.pack(fill="both", expand=True, padx=20, pady=(20, 10))
        
        roster_header = tk.Frame(roster_card, bg=self.colors['info'], height=50)
        roster_header.pack(fill="x")
        roster_header.pack_propagate(False)
        
        tk.Label(roster_header, text="📝 Class Roster - Live Attendance", 
                font=("Segoe UI", 13, "bold"), 
                bg=self.colors['info'], 
                fg='white').pack(side=tk.LEFT, padx=20, pady=12)
        
        roster_body = tk.Frame(roster_card, bg='white')
        roster_body.pack(fill="both", expand=True, padx=20, pady=20)
        
        roster_tree_frame = tk.Frame(roster_body, bg='white')
        roster_tree_frame.pack(fill="both", expand=True)
        
        roster_scrollbar = ttk.Scrollbar(roster_tree_frame)
        roster_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configure treeview style
        style = ttk.Style()
        style.configure('Roster.Treeview', 
                       font=('Segoe UI', 10),
                       rowheight=30,
                       background='white',
                       fieldbackground='white')
        style.configure('Roster.Treeview.Heading',
                       font=('Segoe UI', 11, 'bold'),
                       background=self.colors['bg_secondary'],
                       foreground=self.colors['text_primary'])
        
        self.roster_table = ttk.Treeview(roster_tree_frame, columns=("name", "time"), 
                                        height=12, yscrollcommand=roster_scrollbar.set, 
                                        show="headings", style='Roster.Treeview')
        roster_scrollbar.config(command=self.roster_table.yview)
        self.roster_table.column("name", anchor=tk.W, width=300)
        self.roster_table.column("time", anchor=tk.CENTER, width=200)
        self.roster_table.heading("name", text="👤 Student Name")
        self.roster_table.heading("time", text="🕐 Arrival Time")
        self.roster_table.tag_configure("pending", foreground=self.colors['text_secondary'])
        self.roster_table.tag_configure("present_on_time", foreground=self.colors['success'], 
                                       font=("Segoe UI", 10, "bold"))
        self.roster_table.tag_configure("present_late", foreground=self.colors['warning'], 
                                       font=("Segoe UI", 10, "bold"))
        self.roster_table.tag_configure("absent", foreground=self.colors['danger'], 
                                       font=("Segoe UI", 10, "bold"))
        self.roster_table.pack(fill="both", expand=True, side=tk.LEFT)
        
        # Logs Section - Modern Card
        log_card = tk.Frame(roster_tab, bg='white', relief='flat')
        log_card.pack(fill="x", padx=20, pady=(10, 20))
        
        log_header = tk.Frame(log_card, bg=self.colors['dark'], height=50)
        log_header.pack(fill="x")
        log_header.pack_propagate(False)
        
        tk.Label(log_header, text="📜 Activity Logs", 
                font=("Segoe UI", 12, "bold"), 
                bg=self.colors['dark'], 
                fg='white').pack(side=tk.LEFT, padx=20, pady=12)
        
        log_body = tk.Frame(log_card, bg='white')
        log_body.pack(fill="x", padx=20, pady=20)
        
        log_scrollbar = ttk.Scrollbar(log_body)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(log_body, height=6, state="disabled", 
                               yscrollcommand=log_scrollbar.set, 
                               font=("Consolas", 9), 
                               bg=self.colors['bg_secondary'],
                               bd=0, relief='flat')
        log_scrollbar.config(command=self.log_text.yview)
        self.log_text.pack(fill="both", expand=True, side=tk.LEFT)
        
        # Export buttons
        button_frame = tk.Frame(log_body, bg='white')
        button_frame.pack(pady=(15, 0), fill="x")
        
        self._create_styled_button(button_frame, "📄 Export CSV", self._export_csv, 
                                   self.colors['success'], 'white').pack(side=tk.LEFT, padx=(0, 8), ipady=8, ipadx=15)
        self._create_styled_button(button_frame, "📊 Export Excel", self._export_excel, 
                                   self.colors['warning'], 'white').pack(side=tk.LEFT, padx=8, ipady=8, ipadx=15)
        self._create_styled_button(button_frame, "📑 Detailed Report", self._export_csv_detailed, 
                                   self.colors['info'], 'white').pack(side=tk.LEFT, padx=8, ipady=8, ipadx=15)
        
        # Tab 3: Teachers Management
        teachers_tab = tk.Frame(notebook)
        notebook.add(teachers_tab, text="Teachers")
        
        # Add Teacher Section
        add_teacher_frame = tk.LabelFrame(teachers_tab, text="Add Teacher", font=("Arial", 11, "bold"))
        add_teacher_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(add_teacher_frame, text="Teacher Name:", font=("Arial", 10)).pack(**pad, anchor="w")
        self.teacher_name_var = tk.StringVar()
        tk.Entry(add_teacher_frame, textvariable=self.teacher_name_var, width=30, font=("Arial", 10)).pack(fill="x", **pad)
        
        tk.Button(add_teacher_frame, text="Choose Images", command=self._select_teacher_images, width=25).pack(**pad, fill="x")
        tk.Button(add_teacher_frame, text="Capture from Camera", command=self._capture_teacher_camera, width=25, bg="#2196F3", fg="white").pack(**pad, fill="x")
        tk.Button(add_teacher_frame, text="Save Teacher", command=self._save_teacher, width=25, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(**pad, fill="x")
        
        self.selected_teacher_images: List[str] = []
        self.captured_teacher_encodings: List[np.ndarray] = []
        
        # Teacher List Section
        teacher_list_frame = tk.LabelFrame(teachers_tab, text="Teachers List", font=("Arial", 11, "bold"))
        teacher_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        tk.Label(teacher_list_frame, text="All Teachers:", font=("Arial", 9)).pack(**pad, anchor="w")
        
        teacher_listbox_frame = tk.Frame(teacher_list_frame)
        teacher_listbox_frame.pack(fill="both", expand=True, **pad)
        
        teacher_scrollbar = ttk.Scrollbar(teacher_listbox_frame)
        teacher_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.teacher_listbox = tk.Listbox(teacher_listbox_frame, height=8, yscrollcommand=teacher_scrollbar.set, font=("Arial", 10))
        teacher_scrollbar.config(command=self.teacher_listbox.yview)
        self.teacher_listbox.pack(fill="both", expand=True, side=tk.LEFT)
        
        teacher_btn_frame = tk.Frame(teacher_list_frame)
        teacher_btn_frame.pack(fill="x", **pad)
        tk.Button(teacher_btn_frame, text="Edit", command=self._edit_teacher, width=12).pack(side=tk.LEFT, padx=2)
        tk.Button(teacher_btn_frame, text="Refresh", command=self._refresh_teacher_list, width=12).pack(side=tk.LEFT, padx=2)
        tk.Button(teacher_btn_frame, text="Delete", command=self._delete_teacher, width=12, bg="#f44336", fg="white").pack(side=tk.LEFT, padx=2)
        
        # Tab 4: Contact Information
        contact_tab = tk.Frame(notebook)
        notebook.add(contact_tab, text="Contact Info")
        
        # Student Search
        contact_search_frame = tk.LabelFrame(contact_tab, text="Find Student", font=("Arial", 10, "bold"))
        contact_search_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(contact_search_frame, text="Search Student:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.contact_search_var = tk.StringVar()
        contact_search_entry = tk.Entry(contact_search_frame, textvariable=self.contact_search_var, font=("Arial", 10), width=30)
        contact_search_entry.pack(side=tk.LEFT, padx=5)
        
        def search_and_show_contact():
            search_term = self.contact_search_var.get().strip()
            students = self.db.conn.execute("SELECT id, name FROM students WHERE name LIKE ? ORDER BY name", (f"%{search_term}%",)).fetchall()
            contact_listbox.delete(0, tk.END)
            for sid, sname in students:
                contact_listbox.insert(tk.END, f"{sname} (ID: {sid})")
        
        tk.Button(contact_search_frame, text="Search", command=search_and_show_contact).pack(side=tk.LEFT, padx=5)
        contact_search_entry.bind("<Return>", lambda e: search_and_show_contact())
        
        # Student list
        contact_list_frame = tk.Frame(contact_tab)
        contact_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        contact_scrollbar = ttk.Scrollbar(contact_list_frame)
        contact_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        contact_listbox = tk.Listbox(contact_list_frame, height=8, yscrollcommand=contact_scrollbar.set, font=("Arial", 9))
        contact_scrollbar.config(command=contact_listbox.yview)
        contact_listbox.pack(fill="both", expand=True, side=tk.LEFT)
        contact_listbox.bind("<<ListboxSelect>>", lambda e: self._show_contact_info(contact_listbox))
        
        # Contact Information Display
        info_frame = tk.LabelFrame(contact_tab, text="Contact Details", font=("Arial", 10, "bold"))
        info_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.contact_info_text = tk.Text(info_frame, height=10, state="disabled", font=("Courier", 9))
        self.contact_info_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Buttons for Edit/Add Contact Info
        button_frame = tk.Frame(contact_tab)
        button_frame.pack(fill="x", padx=10, pady=5)
        tk.Button(button_frame, text="Edit Contact Info", width=15, font=("Arial", 9),
                 command=lambda: self._edit_contact_info_dialog(contact_listbox)).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Refresh", width=15, font=("Arial", 9),
                 command=lambda: self._refresh_contact_list(contact_listbox)).pack(side=tk.LEFT, padx=5)
        
        # Store reference to contact_listbox for later use
        self.contact_listbox_ref = contact_listbox
        
        # Tab 5: Reports
        reports_tab = tk.Frame(notebook)
        notebook.add(reports_tab, text="Reports")
        
        # Date Range Section
        date_frame = tk.LabelFrame(reports_tab, text="Select Date Range", font=("Arial", 10, "bold"))
        date_frame.pack(fill="x", padx=10, pady=10)
        
        date_frame_inner = tk.Frame(date_frame)
        date_frame_inner.pack(fill="x", padx=10, pady=10)
        
        tk.Label(date_frame_inner, text="From Date (YYYY-MM-DD):", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.report_from_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"))
        tk.Entry(date_frame_inner, textvariable=self.report_from_date_var, width=15, font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        
        tk.Label(date_frame_inner, text="To Date (YYYY-MM-DD):", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        self.report_to_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(date_frame_inner, textvariable=self.report_to_date_var, width=15, font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        
        # Buttons for quick date ranges
        button_frame = tk.Frame(date_frame)
        button_frame.pack(fill="x", padx=10, pady=5)
        
        def set_this_week():
            today = datetime.now()
            week_start = today - timedelta(days=today.weekday())
            self.report_from_date_var.set(week_start.strftime("%Y-%m-%d"))
            self.report_to_date_var.set(today.strftime("%Y-%m-%d"))
        
        def set_this_month():
            today = datetime.now()
            month_start = today.replace(day=1)
            self.report_from_date_var.set(month_start.strftime("%Y-%m-%d"))
            self.report_to_date_var.set(today.strftime("%Y-%m-%d"))
        
        def set_last_month():
            today = datetime.now()
            first_of_this_month = today.replace(day=1)
            last_of_last_month = first_of_this_month - timedelta(days=1)
            first_of_last_month = last_of_last_month.replace(day=1)
            self.report_from_date_var.set(first_of_last_month.strftime("%Y-%m-%d"))
            self.report_to_date_var.set(last_of_last_month.strftime("%Y-%m-%d"))
        
        tk.Button(button_frame, text="This Week", command=set_this_week, width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(button_frame, text="This Month", command=set_this_month, width=12).pack(side=tk.LEFT, padx=3)
        tk.Button(button_frame, text="Last Month", command=set_last_month, width=12).pack(side=tk.LEFT, padx=3)
        
        # Report Display Section
        report_display_frame = tk.LabelFrame(reports_tab, text="Attendance Statistics", font=("Arial", 10, "bold"))
        report_display_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Treeview for statistics
        report_tree_frame = tk.Frame(report_display_frame)
        report_tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        report_scrollbar = ttk.Scrollbar(report_tree_frame)
        report_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.report_table = ttk.Treeview(report_tree_frame, columns=("student", "total", "present", "late", "absent", "rate"),
                                        height=15, yscrollcommand=report_scrollbar.set, show="headings")
        report_scrollbar.config(command=self.report_table.yview)
        
        self.report_table.column("student", anchor=tk.W, width=150)
        self.report_table.column("total", anchor=tk.CENTER, width=70)
        self.report_table.column("present", anchor=tk.CENTER, width=70)
        self.report_table.column("late", anchor=tk.CENTER, width=70)
        self.report_table.column("absent", anchor=tk.CENTER, width=70)
        self.report_table.column("rate", anchor=tk.CENTER, width=80)
        
        self.report_table.heading("student", text="Student Name")
        self.report_table.heading("total", text="Total Classes")
        self.report_table.heading("present", text="Present")
        self.report_table.heading("late", text="Late")
        self.report_table.heading("absent", text="Absent")
        self.report_table.heading("rate", text="Rate %")
        
        self.report_table.tag_configure("good", foreground="green", font=("Arial", 10))
        self.report_table.tag_configure("warning", foreground="orange", font=("Arial", 10))
        self.report_table.tag_configure("bad", foreground="red", font=("Arial", 10))
        
        self.report_table.pack(fill="both", expand=True)
        
        # Action Buttons
        report_button_frame = tk.Frame(reports_tab)
        report_button_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(report_button_frame, text="Generate Report", command=self._generate_attendance_report, 
             width=20, bg="#2196F3", fg="white", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(report_button_frame, text="Export Report as CSV", command=self._export_report_csv,
             width=20, bg="#4CAF50", fg="white").pack(side=tk.LEFT, padx=5)

        # PDF Report Card Section
        pdf_frame = tk.LabelFrame(reports_tab, text="Student PDF Report Card", font=("Arial", 10, "bold"))
        pdf_frame.pack(fill="x", padx=10, pady=10)
        
        pdf_frame_inner = tk.Frame(pdf_frame)
        pdf_frame_inner.pack(fill="x", padx=10, pady=10)
        
        tk.Label(pdf_frame_inner, text="Select Student:", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        
        self.pdf_student_var = tk.StringVar()
        self.pdf_student_dropdown = tk.OptionMenu(pdf_frame_inner, self.pdf_student_var, "")
        self.pdf_student_dropdown.config(width=30, font=("Arial", 9))
        self.pdf_student_dropdown.pack(side=tk.LEFT, padx=5)
        
        tk.Button(pdf_frame_inner, text="📄 Generate PDF Report Card", command=self._generate_student_pdf_report,
                  bg="#E74C3C", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=10)

        # Tab 6: Attendance History
        history_tab = tk.Frame(notebook, bg=self.colors['bg_secondary'])
        notebook.add(history_tab, text="📜 History")
        
        # Store history tab index
        self.history_tab_index = 6
        
        # History Header
        history_header = self._create_gradient_frame(history_tab, height=70, 
                                                     color1=self.colors['primary'], 
                                                     color2=self.colors['primary_dark'])
        history_header.pack(fill="x", padx=20, pady=(20, 0))
        history_header.pack_propagate(False)
        
        header_content = tk.Frame(history_header, bg=self.colors['primary'])
        header_content.pack(expand=True)
        
        tk.Label(header_content, text="📅 View Attendance History by Date", 
                font=("Segoe UI", 16, "bold"), 
                bg=self.colors['primary'], 
                fg='white').pack(anchor='w')
        
        tk.Label(header_content, text="Select a date to view all classes and attendance details", 
                font=("Segoe UI", 10), 
                bg=self.colors['primary'], 
                fg='#f5f5f5').pack(anchor='w', pady=(5, 0))
        
        # Date Selection Card
        date_card = tk.Frame(history_tab, bg=self.colors['bg_card'], relief='flat')
        date_card.pack(fill="x", padx=20, pady=20)
        
        date_card_body = tk.Frame(date_card, bg=self.colors['bg_card'])
        date_card_body.pack(fill="x", padx=20, pady=20)
        
        # Date input with quick access buttons
        date_row = tk.Frame(date_card_body, bg=self.colors['bg_card'])
        date_row.pack(fill="x", pady=(0, 15))
        
        tk.Label(date_row, text="📅 Select Date:", 
                font=("Segoe UI", 11, "bold"), 
                bg=self.colors['bg_card'], 
                fg=self.colors['text_primary']).pack(side=tk.LEFT, padx=(0, 10))
        
        self.history_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        date_entry = tk.Entry(date_row, textvariable=self.history_date_var, 
                             font=("Segoe UI", 11), width=15, bd=1, relief='solid')
        date_entry.pack(side=tk.LEFT, ipady=8, padx=(0, 10))
        
        # Quick date buttons
        self._create_styled_button(date_row, "Today", 
                                   lambda: self.history_date_var.set(datetime.now().strftime("%Y-%m-%d")), 
                                   self.colors['info'], 'white', width=10).pack(side=tk.LEFT, padx=5, ipady=6)
        
        self._create_styled_button(date_row, "Yesterday", 
                                   lambda: self.history_date_var.set((datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")), 
                                   self.colors['info'], 'white', width=10).pack(side=tk.LEFT, padx=5, ipady=6)
        
        self._create_styled_button(date_row, "🔍 View History", 
                                   self._load_attendance_history, 
                                   self.colors['success'], 'white', width=15).pack(side=tk.LEFT, padx=(20, 0), ipady=8)
        
        # Available dates dropdown
        dates_row = tk.Frame(date_card_body, bg=self.colors['bg_card'])
        dates_row.pack(fill="x", pady=(10, 0))
        
        tk.Label(dates_row, text="📋 Or select from available dates:", 
                font=("Segoe UI", 10), 
                bg=self.colors['bg_card'], 
                fg=self.colors['text_secondary']).pack(side=tk.LEFT, padx=(0, 10))
        
        self.available_dates_var = tk.StringVar()
        self.dates_dropdown = tk.OptionMenu(dates_row, self.available_dates_var, "")
        self.dates_dropdown.config(font=("Segoe UI", 9), bd=1, relief='solid', 
                                   bg='white', activebackground=self.colors['bg_secondary'])
        self.dates_dropdown.pack(side=tk.LEFT, padx=(0, 10))
        
        self._create_styled_button(dates_row, "Load Selected", 
                                   self._load_selected_date, 
                                   self.colors['accent'], 'white', width=12).pack(side=tk.LEFT, ipady=6)
        
        self._create_styled_button(dates_row, "🔄 Refresh Dates", 
                                   self._refresh_available_dates, 
                                   self.colors['info'], 'white', width=12).pack(side=tk.LEFT, padx=(10, 0), ipady=6)
        
        # Load available dates on startup
        self._refresh_available_dates()
        
        # History Display Area
        history_display_card = tk.Frame(history_tab, bg=self.colors['bg_card'], relief='flat')
        history_display_card.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        display_header = tk.Frame(history_display_card, bg=self.colors['dark'], height=45)
        display_header.pack(fill="x")
        display_header.pack_propagate(False)
        
        tk.Label(display_header, text="📊 Classes & Attendance Records", 
                font=("Segoe UI", 12, "bold"), 
                bg=self.colors['dark'], 
                fg='white').pack(side=tk.LEFT, padx=20, pady=10)
        
        display_body = tk.Frame(history_display_card, bg=self.colors['bg_card'])
        display_body.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Treeview for history
        history_tree_frame = tk.Frame(display_body, bg=self.colors['bg_card'])
        history_tree_frame.pack(fill="both", expand=True)
        
        history_scrollbar = ttk.Scrollbar(history_tree_frame)
        history_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        style = ttk.Style()
        style.configure('History.Treeview', 
                       font=('Segoe UI', 10),
                       rowheight=28,
                       background='white',
                       fieldbackground='white')
        style.configure('History.Treeview.Heading',
                       font=('Segoe UI', 11, 'bold'),
                       background=self.colors['bg_secondary'],
                       foreground=self.colors['text_primary'])
        
        self.history_table = ttk.Treeview(history_tree_frame, 
                                         columns=("class", "student", "status", "time", "teacher"),
                                         height=18, 
                                         yscrollcommand=history_scrollbar.set, 
                                         show="headings", 
                                         style='History.Treeview')
        history_scrollbar.config(command=self.history_table.yview)
        
        self.history_table.column("class", anchor=tk.W, width=200)
        self.history_table.column("student", anchor=tk.W, width=200)
        self.history_table.column("status", anchor=tk.CENTER, width=120)
        self.history_table.column("time", anchor=tk.CENTER, width=180)
        self.history_table.column("teacher", anchor=tk.W, width=150)
        
        self.history_table.heading("class", text="📚 Class Name")
        self.history_table.heading("student", text="👤 Student Name")
        self.history_table.heading("status", text="📊 Status")
        self.history_table.heading("time", text="🕐 Time")
        self.history_table.heading("teacher", text="👨‍🏫 Teacher")
        
        # Tags for coloring
        self.history_table.tag_configure("present", foreground=self.colors['success'], 
                                        font=("Segoe UI", 10, "bold"))
        self.history_table.tag_configure("late", foreground=self.colors['warning'], 
                                        font=("Segoe UI", 10, "bold"))
        self.history_table.tag_configure("absent", foreground=self.colors['danger'], 
                                        font=("Segoe UI", 10, "bold"))
        self.history_table.tag_configure("header_row", background=self.colors['bg_secondary'], 
                                        font=("Segoe UI", 11, "bold"))
        
        self.history_table.pack(fill="both", expand=True, side=tk.LEFT)
        
        # Export buttons for history
        export_frame = tk.Frame(display_body, bg=self.colors['bg_card'])
        export_frame.pack(pady=(15, 0), fill="x")
        
        self._create_styled_button(export_frame, "📄 Export to CSV", 
                                   self._export_history_csv, 
                                   self.colors['success'], 'white', width=18).pack(side=tk.LEFT, padx=(0, 8), ipady=8)
        
        self._create_styled_button(export_frame, "📊 Export to Excel", 
                                   self._export_history_excel, 
                                   self.colors['warning'], 'white', width=18).pack(side=tk.LEFT, padx=8, ipady=8)
        
        # Statistics summary
        stats_summary_frame = tk.Frame(date_card_body, bg=self.colors['bg_secondary'], 
                                      bd=1, relief='solid')
        stats_summary_frame.pack(fill="x", pady=(10, 0))
        
        self.history_stats_var = tk.StringVar(value="Select a date and click 'View History' to see attendance records")
        tk.Label(stats_summary_frame, textvariable=self.history_stats_var, 
                font=("Segoe UI", 9), wraplength=800, justify="left", 
                bg=self.colors['bg_secondary'], 
                fg=self.colors['text_primary']).pack(padx=15, pady=12)

        # Tab 7: ⚠️ Warnings
        warnings_tab = tk.Frame(notebook, bg=self.colors['bg_secondary'])
        notebook.add(warnings_tab, text="⚠️ Warnings")
        self.warnings_tab = warnings_tab

        # Warnings Header
        warnings_header = self._create_gradient_frame(warnings_tab, height=70,
                                                       color1='#e74c3c',
                                                       color2='#c0392b')
        warnings_header.pack(fill="x", padx=20, pady=(20, 0))
        warnings_header.pack_propagate(False)

        warn_header_content = tk.Frame(warnings_header, bg='#e74c3c')
        warn_header_content.pack(expand=True)

        tk.Label(warn_header_content, text="⚠️ Student Warnings & Absence Tracking",
                font=("Segoe UI", 16, "bold"),
                bg='#e74c3c',
                fg='white').pack(anchor='w')

        tk.Label(warn_header_content, text="Per-subject warning system — each class tracks absences independently",
                font=("Segoe UI", 10),
                bg='#e74c3c',
                fg='#f5f5f5').pack(anchor='w', pady=(5, 0))

        # Filter Controls Card
        filter_card = tk.Frame(warnings_tab, bg=self.colors['bg_card'], relief='flat')
        filter_card.pack(fill="x", padx=20, pady=15)

        filter_body = tk.Frame(filter_card, bg=self.colors['bg_card'])
        filter_body.pack(fill="x", padx=20, pady=15)

        filter_row = tk.Frame(filter_body, bg=self.colors['bg_card'])
        filter_row.pack(fill="x")

        tk.Label(filter_row, text="📚 Filter by Class:",
                font=("Segoe UI", 11, "bold"),
                bg=self.colors['bg_card'],
                fg=self.colors['text_primary']).pack(side=tk.LEFT, padx=(0, 10))

        self.warning_class_filter_var = tk.StringVar(value="All Classes")
        self.warning_class_dropdown = tk.OptionMenu(filter_row, self.warning_class_filter_var, "All Classes")
        self.warning_class_dropdown.config(font=("Segoe UI", 10), bd=1, relief='solid',
                                           bg='white', activebackground=self.colors['bg_secondary'])
        self.warning_class_dropdown.pack(side=tk.LEFT, padx=(0, 15))

        self._create_styled_button(filter_row, "🔍 Show Warnings",
                                   self._refresh_warnings_tab,
                                   self.colors['danger'], 'white', width=15).pack(side=tk.LEFT, padx=5, ipady=8)

        self._create_styled_button(filter_row, "🔄 Refresh",
                                   self._refresh_warnings_dropdown,
                                   self.colors['info'], 'white', width=12).pack(side=tk.LEFT, padx=5, ipady=8)

        self._create_styled_button(filter_row, "📊 Export Excel",
                                   self._export_warnings_excel,
                                   self.colors['warning'], 'white', width=14).pack(side=tk.LEFT, padx=5, ipady=8)

        # Warning Legend
        legend_frame = tk.Frame(filter_body, bg=self.colors['bg_secondary'], bd=1, relief='solid')
        legend_frame.pack(fill="x", pady=(10, 0))

        legend_inner = tk.Frame(legend_frame, bg=self.colors['bg_secondary'])
        legend_inner.pack(padx=15, pady=8)

        tk.Label(legend_inner, text="📋 Warning Rules:",
                font=("Segoe UI", 9, "bold"),
                bg=self.colors['bg_secondary'],
                fg=self.colors['text_primary']).pack(side=tk.LEFT, padx=(0, 15))

        tk.Label(legend_inner, text="2 Absences → 1st Warning",
                font=("Segoe UI", 9),
                bg=self.colors['bg_secondary'],
                fg='#f39c12').pack(side=tk.LEFT, padx=(0, 15))

        tk.Label(legend_inner, text="3 Absences → 2nd Warning",
                font=("Segoe UI", 9),
                bg=self.colors['bg_secondary'],
                fg='#e67e22').pack(side=tk.LEFT, padx=(0, 15))

        tk.Label(legend_inner, text="4+ Absences → Last Warning",
                font=("Segoe UI", 9),
                bg=self.colors['bg_secondary'],
                fg='#e74c3c').pack(side=tk.LEFT)

        # Warnings Table Card
        warnings_table_card = tk.Frame(warnings_tab, bg=self.colors['bg_card'], relief='flat')
        warnings_table_card.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        warn_table_header = tk.Frame(warnings_table_card, bg=self.colors['danger'], height=45)
        warn_table_header.pack(fill="x")
        warn_table_header.pack_propagate(False)

        tk.Label(warn_table_header, text="📊 Student Warnings by Class",
                font=("Segoe UI", 12, "bold"),
                bg=self.colors['danger'],
                fg='white').pack(side=tk.LEFT, padx=20, pady=10)

        self.warnings_count_label = tk.Label(warn_table_header, text="",
                                              font=("Segoe UI", 10),
                                              bg=self.colors['danger'],
                                              fg='#f5f5f5')
        self.warnings_count_label.pack(side=tk.RIGHT, padx=20, pady=10)

        warn_table_body = tk.Frame(warnings_table_card, bg=self.colors['bg_card'])
        warn_table_body.pack(fill="both", expand=True, padx=20, pady=20)

        warn_tree_frame = tk.Frame(warn_table_body, bg=self.colors['bg_card'])
        warn_tree_frame.pack(fill="both", expand=True)

        warn_scrollbar = ttk.Scrollbar(warn_tree_frame)
        warn_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        style.configure('Warnings.Treeview',
                        font=('Segoe UI', 10),
                        rowheight=30,
                        background='white',
                        fieldbackground='white')
        style.configure('Warnings.Treeview.Heading',
                        font=('Segoe UI', 11, 'bold'),
                        background=self.colors['bg_secondary'],
                        foreground=self.colors['text_primary'])

        self.warnings_table = ttk.Treeview(warn_tree_frame,
                                           columns=("student", "class", "total", "attended", "absent", "warning"),
                                           height=18,
                                           yscrollcommand=warn_scrollbar.set,
                                           show="headings",
                                           style='Warnings.Treeview')
        warn_scrollbar.config(command=self.warnings_table.yview)

        self.warnings_table.column("student", anchor=tk.W, width=200)
        self.warnings_table.column("class", anchor=tk.W, width=180)
        self.warnings_table.column("total", anchor=tk.CENTER, width=120)
        self.warnings_table.column("attended", anchor=tk.CENTER, width=120)
        self.warnings_table.column("absent", anchor=tk.CENTER, width=100)
        self.warnings_table.column("warning", anchor=tk.CENTER, width=150)

        self.warnings_table.heading("student", text="👤 Student Name")
        self.warnings_table.heading("class", text="📚 Class Name")
        self.warnings_table.heading("total", text="📅 Total Lectures")
        self.warnings_table.heading("attended", text="✓ Attended")
        self.warnings_table.heading("absent", text="❌ Absent")
        self.warnings_table.heading("warning", text="⚠️ Warning Level")

        self.warnings_table.tag_configure("no_warning", foreground=self.colors['success'],
                                          font=("Segoe UI", 10))
        self.warnings_table.tag_configure("warning_1st", foreground='#f39c12',
                                          font=("Segoe UI", 10, "bold"))
        self.warnings_table.tag_configure("warning_2nd", foreground='#e67e22',
                                          font=("Segoe UI", 10, "bold"))
        self.warnings_table.tag_configure("warning_last", foreground='#e74c3c',
                                          font=("Segoe UI", 10, "bold"))

        self.warnings_table.pack(fill="both", expand=True, side=tk.LEFT)

        # Initialize warnings dropdown
        self._refresh_warnings_dropdown()

        # Tab 8: 📈 Dashboard
        dashboard_tab = tk.Frame(notebook, bg=self.colors['bg_secondary'])
        notebook.add(dashboard_tab, text="📈 Dashboard")
        self.dashboard_tab = dashboard_tab
        
        # Dashboard Header
        dash_header = self._create_gradient_frame(dashboard_tab, height=70,
                                                       color1='#8E44AD',
                                                       color2='#9B59B6')
        dash_header.pack(fill="x", padx=20, pady=(20, 0))
        dash_header.pack_propagate(False)

        dash_header_content = tk.Frame(dash_header, bg='#8E44AD')
        dash_header_content.pack(expand=True)

        tk.Label(dash_header_content, text="📈 Attendance Analytics & Dashboard",
                font=("Segoe UI", 16, "bold"),
                bg='#8E44AD',
                fg='white').pack(anchor='w')
                
        # Top Controls
        dash_controls = tk.Frame(dashboard_tab, bg=self.colors['bg_secondary'])
        dash_controls.pack(fill="x", padx=20, pady=10)
        
        self._create_styled_button(dash_controls, "🔄 Refresh Charts", 
                                   self._refresh_dashboard, 
                                   self.colors['primary'], 'white', width=15).pack(side=tk.LEFT)
                                   
        # Charts Container
        self.charts_frame = tk.Frame(dashboard_tab, bg=self.colors['bg_card'], bd=1, relief="solid")
        self.charts_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Tab 9: Settings
        settings_tab = tk.Frame(notebook)
        notebook.add(settings_tab, text="Settings")

        # Backup & Restore
        backup_frame = tk.LabelFrame(settings_tab, text="Backup & Restore", font=("Arial", 10, "bold"))
        backup_frame.pack(fill="x", padx=10, pady=10)

        tk.Button(backup_frame, text="Backup Database", width=20, bg="#4CAF50", fg="white",
              command=self._backup_database).pack(side=tk.LEFT, padx=5, pady=8)
        tk.Button(backup_frame, text="Restore Database", width=20, bg="#f44336", fg="white",
              command=self._restore_database).pack(side=tk.LEFT, padx=5, pady=8)

        # Face tolerance and camera settings
        tolerance_frame = tk.LabelFrame(settings_tab, text="Face & Camera", font=("Arial", 10, "bold"))
        tolerance_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(tolerance_frame, text="Face match tolerance (lower = stricter):", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 2))
        self.encoding_tolerance_var = tk.DoubleVar(value=ENCODING_TOLERANCE)
        tk.Scale(tolerance_frame, variable=self.encoding_tolerance_var, from_=0.3, to=0.7, resolution=0.01,
             orient=tk.HORIZONTAL, length=280).pack(anchor="w", padx=12, pady=(0, 8))

        cam_row = tk.Frame(tolerance_frame)
        cam_row.pack(fill="x", padx=10, pady=4)
        tk.Label(cam_row, text="Default Camera Index:", font=("Arial", 9)).pack(side=tk.LEFT)
        tk.Entry(cam_row, textvariable=self.camera_index_var, width=6, font=("Arial", 9)).pack(side=tk.LEFT, padx=6)
        tk.Button(cam_row, text="Test", width=10, command=self._test_camera).pack(side=tk.LEFT, padx=6)

        # Theme settings
        theme_frame = tk.LabelFrame(settings_tab, text="Theme", font=("Arial", 10, "bold"))
        theme_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(theme_frame, text="Choose color theme:", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 2))
        self.theme_var = tk.StringVar(value="light")
        theme_options = ["light", "dark", "blue"]
        theme_row = tk.Frame(theme_frame)
        theme_row.pack(fill="x", padx=10, pady=(0, 8))
        theme_menu = tk.OptionMenu(theme_row, self.theme_var, *theme_options)
        theme_menu.pack(side=tk.LEFT)
        tk.Button(theme_row, text="Apply Theme", command=self._apply_theme, width=14).pack(side=tk.LEFT, padx=6)

        # Apply settings button
        tk.Button(settings_tab, text="Save Settings", width=18, bg="#2196F3", fg="white",
              command=self._apply_settings).pack(padx=12, pady=12, anchor="w")

        # Initialize lists
        self._refresh_teacher_list()
        
        # Initialize student list
        self._refresh_student_list()
        self._refresh_class_list()
        self._refresh_contact_list(contact_listbox)
        self._detect_camera_auto()

    def _refresh_class_list(self):
        """Refresh the class listbox with all classes"""
        self.class_listbox.delete(0, tk.END)
        self.class_list_items = []  # keep mapping for selection → class/session

        classes = self.db.conn.execute("SELECT id, name FROM classes ORDER BY name").fetchall()
        for class_id, name in classes:
            sessions = self.db.conn.execute(
                "SELECT session_number, session_date FROM class_sessions WHERE class_id = ? ORDER BY session_number",
                (class_id,),
            ).fetchall()

            if sessions:
                for session_number, session_date in sessions:
                    label = f"{name}{session_number}"  # e.g., Math1, Math2
                    self.class_listbox.insert(tk.END, label)
                    self.class_list_items.append(
                        {"class_id": class_id, "class_name": name, "session_number": session_number, "session_date": session_date}
                    )
            else:
                # No sessions yet; show plain class name
                self.class_listbox.insert(tk.END, name)
                self.class_list_items.append(
                    {"class_id": class_id, "class_name": name, "session_number": None, "session_date": None}
                )
    
    def _on_class_selected(self, event=None):
        """Handle class selection - show attendance history"""
        selection = self.class_listbox.curselection()
        if not selection:
            return
        
        # Map selection to class and (optional) session
        if not hasattr(self, "class_list_items") or selection[0] >= len(self.class_list_items):
            return

        sel = self.class_list_items[selection[0]]
        class_id = sel["class_id"]
        class_name = sel["class_name"]
        session_number = sel.get("session_number") or 0
        session_date = sel.get("session_date")
        
        # Switch to History tab
        if hasattr(self, 'main_notebook'):
            # Find the History tab index (it's tab 6: Dashboard, Classes, Roster, Teachers, Contact, Reports, History)
            history_tab_index = 6  # 0-indexed
            self.main_notebook.select(history_tab_index)
            
            # Use session date if known, otherwise today
            target_date = session_date or datetime.now().strftime("%Y-%m-%d")
            self.history_date_var.set(target_date)

            # Load history for that date
            self._load_attendance_history()

            # Scroll to the matching class/session header
            display_name = f"{class_name} (Session #{session_number})" if session_number else class_name
            self._scroll_to_class(display_name)

            self._log(f"📋 Showing attendance history for '{class_name}' on {target_date}")
    
    def _scroll_to_class(self, class_name: str):
        """Scroll to and highlight a specific class in history table"""
        # Find items matching this class name
        for item in self.history_table.get_children():
            values = self.history_table.item(item)["values"]
            if values and len(values) > 0:
                # Check if this is a header row for our class
                if class_name in str(values[0]):
                    self.history_table.see(item)
                    self.history_table.selection_set(item)
                    break
    
    def _detect_camera_auto(self):
        """Auto-detect which camera index works"""
        detected_index = None
        for idx in range(3):  # Try 0, 1, 2
            cap = cv2.VideoCapture(idx)
            if cap.isOpened():
                cap.release()
                detected_index = idx
                break
        
        if detected_index is not None:
            self.camera_index_var.set(str(detected_index))
            self._log(f"✓ Camera detected at index {detected_index}")
        else:
            self._log(f"❌ No camera detected! Try plugging in a camera.")
        
        # Refresh dashboard
        self._refresh_dashboard()
    
    def _refresh_dashboard(self) -> None:
        """Refresh dashboard statistics"""
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Get today's attendance
        current_attendance = self.db.conn.execute(
            "SELECT COUNT(*) FROM attendance WHERE timestamp LIKE ?",
            (f"{today}%",)
        ).fetchone()[0] if hasattr(self, 'current_class_id') else 0
        
        # Count present, late, absent
        if hasattr(self, 'current_class_id'):
            class_id = self.current_class_id
            attendance = self.db.attendance_for_class(class_id)
            absent = self.db.absent_for_class(class_id)
            
            present_on_time = 0
            present_late = 0
            
            if hasattr(self, 'late_cutoff'):
                late_cutoff = self.late_cutoff
                for name, ts_str in attendance:
                    ts = datetime.fromisoformat(ts_str)
                    if ts <= late_cutoff:
                        present_on_time += 1
                    else:
                        present_late += 1
            else:
                present_on_time = len(attendance)
            
            absent_count = len(absent)
        else:
            present_on_time = 0
            present_late = 0
            absent_count = 0
        
        # Count teachers today
        today_teachers = self.db.conn.execute(
            "SELECT COUNT(DISTINCT teacher_id) FROM teacher_attendance WHERE timestamp LIKE ?",
            (f"{today}%",)
        ).fetchone()[0]
        
        # Count unknown visitors today
        today_visitors = self.db.conn.execute(
            "SELECT COUNT(*) FROM unknown_visitors WHERE timestamp LIKE ?",
            (f"{today}%",)
        ).fetchone()[0]
        
        # Count total classes
        total_classes = self.db.conn.execute("SELECT COUNT(*) FROM classes").fetchone()[0]
        
        # Update labels with just numbers (no text)
        self.stat_present_label.config(text=str(present_on_time))
        self.stat_late_label.config(text=str(present_late))
        self.stat_absent_label.config(text=str(absent_count))
        self.stat_teachers_label.config(text=str(today_teachers))
        self.stat_visitors_label.config(text=str(today_visitors))
        if hasattr(self, 'stat_classes_label'):
            self.stat_classes_label.config(text=str(total_classes))


    def _select_images(self) -> None:
        files = filedialog.askopenfilenames(title="Choose face images", filetypes=[("Images", "*.jpg;*.jpeg;*.png")])
        if files:
            self.selected_images = list(files)
            self._log(f"Selected {len(files)} image(s).")
    
    def _refresh_student_list(self):
        """Refresh the student listbox with all students"""
        self.student_listbox.delete(0, tk.END)
        students = self.db.conn.execute("SELECT id, name FROM students ORDER BY name").fetchall()
        for student_id, name in students:
            self.student_listbox.insert(tk.END, f"{name}")
        self.student_info_var.set("Select a student")
        
        # Also refresh PDF student dropdown
        if hasattr(self, 'pdf_student_dropdown'):
            menu = self.pdf_student_dropdown["menu"]
            menu.delete(0, "end")
            for sid, sname in students:
                val = f"{sname} #{sid}"
                menu.add_command(label=val, command=tk._setit(self.pdf_student_var, val))
            if students:
                self.pdf_student_var.set(f"{students[0][1]} #{students[0][0]}")
            else:
                self.pdf_student_var.set("")
    
    def _search_students(self) -> None:
        """Search students by name"""
        search_term = self.student_search_var.get().strip().lower()
        self.student_listbox.delete(0, tk.END)
        
        students = self.db.conn.execute("SELECT id, name FROM students ORDER BY name").fetchall()
        
        if not search_term:
            # Show all students
            for student_id, name in students:
                self.student_listbox.insert(tk.END, f"{name}")
        else:
            # Show filtered students
            for student_id, name in students:
                if search_term in name.lower():
                    self.student_listbox.insert(tk.END, f"{name}")
        
        self.student_info_var.set("Select a student")
    
    def _refresh_teacher_list(self):
        """Refresh the teacher listbox with all teachers"""
        if hasattr(self, 'teacher_listbox'):
            self.teacher_listbox.delete(0, tk.END)
            teachers = self.db.list_teachers()
            for teacher_id, name in teachers:
                self.teacher_listbox.insert(tk.END, f"{name}")
            if teachers:
                self._log(f"✓ {len(teachers)} teacher(s) loaded")
    
    def _select_teacher_images(self) -> None:
        files = filedialog.askopenfilenames(title="Choose teacher face images", filetypes=[("Images", "*.jpg;*.jpeg;*.png")])
        if files:
            self.selected_teacher_images = list(files)
            self._log(f"Selected {len(files)} teacher image(s).")
    
    def _capture_teacher_camera(self) -> None:
        cap = cv2.VideoCapture(self._get_camera_index())
        if not cap.isOpened():
            messagebox.showerror("Camera Error", "Cannot access camera.")
            return
        
        self._log("📷 Camera opened for teacher. SPACE=capture, C=finish")
        captured = 0
        
        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                
                display = frame.copy()
                h, w = display.shape[:2]
                
                cv2.rectangle(display, (0, 0), (w, 80), (0, 0, 0), -1)
                cv2.rectangle(display, (0, h-80), (w, h), (0, 0, 0), -1)
                
                cv2.putText(display, "Capture Teacher Face", (15, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                cv2.putText(display, f"Captured: {captured} | SPACE=capture | C=finish", (15, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 1)
                
                if captured > 0:
                    cv2.putText(display, f"Teacher faces: {captured} ✓", (15, h-20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                else:
                    cv2.putText(display, "Point camera at teacher face and press SPACE", (15, h-20), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 1)
                
                cv2.imshow("Teacher Face Capture - SPACE=capture, C=finish", display)
                key = cv2.waitKey(1) & 0xFF
                
                if key in (ord('c'), ord('C'), 27):
                    break
                elif key == 32:  # space
                    encs = self._encode_frame(frame)
                    if len(encs) == 0:
                        self._log("❌ No face found. Try again.")
                        continue
                    if len(encs) > 1:
                        self._log("⚠️  Multiple faces detected. Capture with one face only.")
                        continue
                    self.captured_teacher_encodings.append(encs[0])
                    captured += 1
                    self._log(f"✓ Captured teacher face #{captured}")
        finally:
            cap.release()
            cv2.destroyAllWindows()
        
        if captured > 0:
            self._log(f"✓ Successfully captured {captured} teacher face(s)")
        else:
            self._log("⚠️  No teacher faces were captured")
    
    def _save_teacher(self) -> None:
        name = self.teacher_name_var.get().strip()
        if not name:
            messagebox.showerror("Missing", "Enter teacher name.")
            return
        
        try:
            encodings: List[np.ndarray] = []
            if self.selected_teacher_images:
                encodings.extend(self._encode_images(self.selected_teacher_images))
            if self.captured_teacher_encodings:
                encodings.extend(self.captured_teacher_encodings)
            
            if not encodings:
                messagebox.showerror("Missing", "Add teacher face via camera or images.")
                return
            
            teacher_id = self.db.add_teacher(name, encodings)
            self._log(f"✓ Added teacher {name} (id={teacher_id}) with {len(encodings)} encoding(s).")
            
            self.teacher_name_var.set("")
            self.selected_teacher_images = []
            self.captured_teacher_encodings = []
            self._refresh_teacher_list()
            messagebox.showinfo("Success", f"Teacher {name} added successfully!")
        except Exception as exc:
            messagebox.showerror("Error", str(exc))
    
    def _delete_teacher(self) -> None:
        selection = self.teacher_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a teacher to delete")
            return
        
        teachers = self.db.list_teachers()
        if selection[0] < len(teachers):
            teacher_id, name = teachers[selection[0]]
            if messagebox.askyesno("Confirm", f"Delete teacher '{name}'?"):
                self.db.delete_teacher(teacher_id)
                self._refresh_teacher_list()
                self._log(f"✓ Teacher '{name}' deleted")
                messagebox.showinfo("Success", f"Teacher '{name}' deleted")
    
    def _edit_teacher(self) -> None:
        """Edit selected teacher - assign to classes"""
        selection = self.teacher_listbox.curselection()
        if not selection:
            messagebox.showwarning("Error", "Select a teacher to edit")
            return
        
        teachers = self.db.list_teachers()
        if selection[0] >= len(teachers):
            return
        
        teacher_id, teacher_name = teachers[selection[0]]
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Teacher: {teacher_name}")
        dialog.geometry("450x400")
        dialog.resizable(False, False)
        
        pad = {"padx": 10, "pady": 8}
        
        # Teacher Name Section
        name_frame = tk.LabelFrame(dialog, text="Teacher Name", font=("Arial", 10, "bold"))
        name_frame.pack(fill="x", **pad)
        
        name_var = tk.StringVar(value=teacher_name)
        tk.Entry(name_frame, textvariable=name_var, font=("Arial", 11), width=40).pack(fill="x", **pad)
        
        # Classes Assignment
        classes_frame_label = tk.LabelFrame(dialog, text="Assign to Classes", font=("Arial", 10, "bold"))
        classes_frame_label.pack(fill="both", expand=True, **pad)
        
        # Load classes
        all_classes = self.db.list_classes()
        current_class_ids = set(self.db.get_teacher_classes(teacher_id))
        
        class_vars = {}
        
        if all_classes:
            tk.Label(classes_frame_label, text="Select Classes:", font=("Arial", 9)).pack(**pad, anchor="w")
            for cid, cname in all_classes:
                var = tk.BooleanVar(value=(cid in current_class_ids))
                tk.Checkbutton(classes_frame_label, text=f"  {cname}", variable=var, font=("Arial", 10), anchor="w").pack(fill="x", **pad)
                class_vars[cid] = var
        else:
            tk.Label(classes_frame_label, text="No classes created yet!", font=("Arial", 9), fg="red").pack(**pad)
        
        # Buttons
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill="x", **pad)
        
        def save_edit():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showerror("Error", "Teacher name cannot be empty")
                return
            
            # Get selected classes
            selected_ids = [cid for cid, var in class_vars.items() if var.get()]
            
            # Update teacher name
            self.db.conn.execute("UPDATE teachers SET name = ? WHERE id = ?", (new_name, teacher_id))
            
            # Update classes
            self.db.update_teacher_classes(teacher_id, selected_ids)
            self.db.conn.commit()
            
            self._log(f"✓ Teacher '{new_name}' updated with {len(selected_ids)} classes")
            self._refresh_teacher_list()
            messagebox.showinfo("Success", f"Teacher '{new_name}' updated!")
            dialog.destroy()
        
        tk.Button(btn_frame, text="Save Changes", command=save_edit, width=20, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=20).pack(side=tk.LEFT, padx=5)
    
    def _on_student_selected(self, event=None):
        """Show selected student's information"""
        selection = self.student_listbox.curselection()
        if not selection:
            return
        
        students = self.db.conn.execute("SELECT id, name FROM students ORDER BY name").fetchall()
        if selection[0] < len(students):
            student_id, name = students[selection[0]]
            
            # Get enrolled classes
            classes = self.db.conn.execute("""
                SELECT c.name FROM classes c
                JOIN class_students cs ON c.id = cs.class_id
                WHERE cs.student_id = ?
            """, (student_id,)).fetchall()
            
            class_names = ", ".join([c[0] for c in classes]) if classes else "No classes"
            
            info_text = f"Name: {name}\nClasses: {class_names}"
            self.student_info_var.set(info_text)
    
    def _show_contact_info(self, contact_listbox) -> None:
        """Show contact information for selected student"""
        selection = contact_listbox.curselection()
        if not selection:
            return
        
        selected_text = contact_listbox.get(selection[0])
        # Extract student ID from "Name (ID: X)"
        try:
            student_id = int(selected_text.split("ID: ")[1].rstrip(")"))
        except:
            return
        
        # Get student name
        student_name = self.db.conn.execute("SELECT name FROM students WHERE id = ?", (student_id,)).fetchone()[0]
        
        # Get contact info
        contact = self.db.conn.execute("SELECT * FROM contact_info WHERE student_id = ?", (student_id,)).fetchone()
        
        # Display info
        self.contact_info_text.config(state="normal")
        self.contact_info_text.delete("1.0", tk.END)
        
        info = f"📋 Student Information\n{'='*40}\n"
        info += f"Name: {student_name}\n"
        info += f"Student ID: {student_id}\n\n"
        
        if contact:
            info += f"📱 Contact Information\n{'='*40}\n"
            info += f"Phone: {contact[2] or 'N/A'}\n"
            info += f"Email: {contact[3] or 'N/A'}\n"
            info += f"Address: {contact[4] or 'N/A'}\n\n"
            info += f"👨‍👩‍👧 Guardian Information\n{'='*40}\n"
            info += f"Guardian Name: {contact[5] or 'N/A'}\n"
            info += f"Guardian Phone: {contact[6] or 'N/A'}\n"
            info += f"Notes: {contact[7] or 'N/A'}\n"
        else:
            info += "No contact information found.\n"
            info += "Click 'Edit' to add contact information.\n"
        
        self.contact_info_text.insert("1.0", info)
        self.contact_info_text.config(state="disabled")
    
    def _edit_selected_student(self):
        """Edit the selected student"""
        selection = self.student_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a student to edit")
            return
        
        students = self.db.conn.execute("SELECT id, name FROM students ORDER BY name").fetchall()
        if selection[0] < len(students):
            student_id, name = students[selection[0]]
            self._open_student_edit_dialog(student_id, name)
    
    def _delete_selected_student(self):
        """Delete the selected student"""
        selection = self.student_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a student to delete")
            return
        
        students = self.db.conn.execute("SELECT id, name FROM students ORDER BY name").fetchall()
        if selection[0] < len(students):
            student_id, name = students[selection[0]]
            if messagebox.askyesno("Confirm", f"Delete student '{name}'? This will remove all attendance records."):
                self.db.delete_student(student_id)
                self._refresh_student_list()
                messagebox.showinfo("Success", f"Student '{name}' deleted successfully")
    
    def _open_student_edit_dialog(self, student_id, current_name):
        """Open dialog to edit student name, face images, and classes"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Student")
        dialog.geometry("550x500")
        dialog.resizable(False, False)
        
        pad = {"padx": 10, "pady": 5}
        
        # ============ NAME SECTION ============
        name_frame = tk.LabelFrame(dialog, text="Name", font=("Arial", 10, "bold"))
        name_frame.pack(fill="x", **pad)
        
        name_var = tk.StringVar(value=current_name)
        name_entry = tk.Entry(name_frame, textvariable=name_var, font=("Arial", 11), width=40)
        name_entry.pack(fill="x", **pad)
        
        # ============ FACE IMAGES SECTION ============
        face_frame = tk.LabelFrame(dialog, text="Face Images/Camera", font=("Arial", 10, "bold"))
        face_frame.pack(fill="x", **pad)
        
        selected_images_var = tk.StringVar(value="No images selected")
        tk.Label(face_frame, textvariable=selected_images_var, font=("Arial", 9), wraplength=450, justify="left").pack(**pad, anchor="w")
        
        def select_new_images():
            files = filedialog.askopenfilenames(title="Choose face images", filetypes=[("Images", "*.jpg;*.jpeg;*.png")])
            if files:
                selected_images_var.set(f"Selected {len(files)} image(s)")
        
        def capture_from_camera():
            dialog.withdraw()
            self.root.update()
            cap = cv2.VideoCapture(self._get_camera_index())
            if not cap.isOpened():
                messagebox.showerror("Camera Error", "Cannot open camera. Check camera connection.")
                dialog.deiconify()
                return
            
            encodings = []
            
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                faces = face_recognition.face_locations(rgb, model="hog")
                
                if faces:
                    for face_loc in faces[:5]:
                        encoding = face_recognition.face_encoding(rgb, face_loc)
                        if len(encoding) > 0:
                            encodings.append(encoding)
                
                h, w = frame.shape[:2]
                cv2.rectangle(frame, (0, 0), (w, 50), (0, 0, 0), -1)
                cv2.putText(frame, f"Captured: {len(encodings)} | Press SPACE to add | C to finish", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                cv2.imshow(f"Capture Face for {current_name}", frame)
                
                key = cv2.waitKey(1) & 0xFF
                if key == ord('c') or key == 27:
                    break
            
            cap.release()
            cv2.destroyAllWindows()
            dialog.deiconify()
            
            if encodings:
                existing_encodings = self.db.get_student_encodings(student_id)
                existing_encodings.extend(encodings)
                enc_json = json.dumps([enc.tolist() for enc in existing_encodings])
                self.db.conn.execute("UPDATE students SET encodings = ? WHERE id = ?", (enc_json, student_id))
                self.db.conn.commit()
                selected_images_var.set(f"Added {len(encodings)} new face(s) from camera")
                messagebox.showinfo("Success", f"Added {len(encodings)} face(s)")
            else:
                messagebox.showwarning("No Faces", "No faces detected")
        
        btn_face_frame = tk.Frame(face_frame)
        btn_face_frame.pack(fill="x", **pad)
        tk.Button(btn_face_frame, text="Add Images", command=select_new_images, width=18).pack(side=tk.LEFT, padx=3)
        tk.Button(btn_face_frame, text="Capture Camera", command=capture_from_camera, width=18, bg="#2196F3", fg="white").pack(side=tk.LEFT, padx=3)
        
        # ============ CLASSES SECTION ============
        classes_frame = tk.LabelFrame(dialog, text="Assign to Classes", font=("Arial", 10, "bold"))
        classes_frame.pack(fill="both", expand=True, **pad)
        
        # Get all classes and student's current classes
        all_classes = self.db.conn.execute("SELECT id, name FROM classes ORDER BY name").fetchall()
        current_classes = self.db.conn.execute("""
            SELECT class_id FROM class_students WHERE student_id = ?
        """, (student_id,)).fetchall()
        current_class_ids = {c[0] for c in current_classes}
        
        class_vars = {}
        
        if all_classes:
            tk.Label(classes_frame, text="Select Classes:", font=("Arial", 9)).pack(**pad, anchor="w")
            for class_id, class_name in all_classes:
                var = tk.BooleanVar(value=(class_id in current_class_ids))
                tk.Checkbutton(classes_frame, text=f"  {class_name}", variable=var, font=("Arial", 10), anchor="w").pack(fill="x", **pad)
                class_vars[class_id] = var
        else:
            tk.Label(classes_frame, text="No classes created yet. Create classes first!", font=("Arial", 9), fg="red").pack(**pad)
        
        # ============ ACTION BUTTONS ============
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill="x", **pad)
        
        def save_changes():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showerror("Error", "Student name cannot be empty")
                return
            
            # Get selected classes from checkbuttons
            selected_class_ids = [class_id for class_id, var in class_vars.items() if var.get()]
            
            self.db.conn.execute("UPDATE students SET name = ? WHERE id = ?", (new_name, student_id))
            self.db.conn.execute("DELETE FROM class_students WHERE student_id = ?", (student_id,))
            for class_id in selected_class_ids:
                self.db.conn.execute("INSERT INTO class_students (student_id, class_id) VALUES (?, ?)",
                                   (student_id, class_id))
            self.db.conn.commit()
            
            self._refresh_student_list()
            self._log(f"✓ Student '{new_name}' updated")
            messagebox.showinfo("Success", f"Student '{new_name}' updated successfully")
            dialog.destroy()
        
        tk.Button(btn_frame, text="Save Changes", command=save_changes, width=20, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=20).pack(side=tk.LEFT, padx=5)



    def _save_student(self) -> None:
        name = self.student_name_var.get().strip()
        if not name:
            messagebox.showerror("Missing", "Enter a student name.")
            return
        try:
            encodings: List[np.ndarray] = []
            if self.selected_images:
                encodings.extend(self._encode_images(self.selected_images))
            if self.captured_encodings:
                encodings.extend(self.captured_encodings)
            if not encodings:
                messagebox.showerror("Missing", "Add a face via camera or images.")
                return
            student_id = self.db.add_student(name, encodings)
            self._log(f"Added student {name} (id={student_id}) with {len(encodings)} encoding(s).")
            
            # Ask for class assignment (optional)
            if messagebox.askyesno("Assign Classes", f"Assign {name} to classes now?"):
                self._assign_student_to_classes(student_id, name)
            else:
                self._log(f"Student {name} added but not assigned to any class yet.")
            
            self.student_name_var.set("")
            self.selected_images = []
            self.captured_encodings = []
            self._refresh_student_list()
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Error", str(exc))

    def _assign_student_to_classes(self, student_id: int, student_name: str) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Assign {student_name} to Classes")
        dialog.geometry("400x300")
        tk.Label(dialog, text=f"Select classes for {student_name}:").pack(pady=4)
        listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, height=10)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        classes = self.db.list_classes()
        for class_id, class_name in classes:
            listbox.insert(tk.END, f"{class_name} #{class_id}")
        def save_assignment() -> None:
            selected = listbox.curselection()
            class_ids = [int(listbox.get(i).split("#")[1]) for i in selected]
            self.db.update_student(student_id, student_name, class_ids)
            if class_ids:
                self._log(f"Assigned {student_name} to {len(class_ids)} class(es).")
            else:
                self._log(f"{student_name} not assigned to any class.")
            dialog.destroy()
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=4)
        tk.Button(button_frame, text="Save", command=save_assignment).pack(side=tk.LEFT, padx=4)
        tk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=4)

    def _encode_images(self, image_paths: List[str]) -> List[np.ndarray]:
        encodings: List[np.ndarray] = []
        for path in image_paths:
            img = face_recognition.load_image_file(path)
            faces = face_recognition.face_encodings(img)
            if not faces:
                raise ValueError(f"No face found in {path}.")
            encodings.extend(faces)
        return encodings

    def _encode_frame(self, frame: np.ndarray) -> List[np.ndarray]:
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        return face_recognition.face_encodings(rgb)

    def _capture_from_camera(self) -> None:
        cap = cv2.VideoCapture(self._get_camera_index())
        if not cap.isOpened():
            messagebox.showerror("Camera Error", "Cannot access camera. Please check:\n1. Camera is connected\n2. No other app is using it\n3. Try different camera index in Start Attendance")
            return
        
        self._log("📷 Camera opened. SPACE=capture face, C=finish")
        captured = 0
        
        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                
                display = frame.copy()
                h, w = display.shape[:2]
                
                # Add background overlay for text
                cv2.rectangle(display, (0, 0), (w, 80), (0, 0, 0), -1)
                cv2.rectangle(display, (0, h-80), (w, h), (0, 0, 0), -1)
                
                # Top instructions
                cv2.putText(display, "Capture Face - Student Enrollment", (15, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                cv2.putText(display, f"Captured: {captured} | Press SPACE to capture | Press C to finish", (15, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 1)
                
                # Bottom status
                if captured > 0:
                    cv2.putText(display, f"Faces captured: {captured} ✓", (15, h-20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                else:
                    cv2.putText(display, "Point camera at face and press SPACE", (15, h-20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 165, 255), 1)
                
                cv2.imshow("Face Capture - Press SPACE to capture, C to finish", display)
                key = cv2.waitKey(1) & 0xFF
                
                if key in (ord('c'), ord('C'), 27):  # c, C, or ESC
                    break
                elif key == 32:  # space
                    encs = self._encode_frame(frame)
                    if len(encs) == 0:
                        self._log("❌ No face found. Try again.")
                        continue
                    if len(encs) > 1:
                        self._log("⚠️  Multiple faces detected. Capture with one face only.")
                        continue
                    self.captured_encodings.append(encs[0])
                    captured += 1
                    self._log(f"✓ Captured face #{captured}")
        finally:
            cap.release()
            cv2.destroyAllWindows()
        
        if captured > 0:
            self._log(f"✓ Successfully captured {captured} face(s) from camera")
        else:
            self._log("⚠️  No faces were captured")


    def _start_attendance_thread(self) -> None:
        try:
            # Keep attendance setup on Tk's main thread; camera processing runs in worker threads.
            self._start_attendance_ui()
        except Exception as exc:
            self._log_exception("Failed to start attendance session", exc)
            self._show_message_async("error", "Attendance Error", "Failed to start attendance. Check attendance_debug.log")

    def _start_attendance_ui(self) -> None:
        class_name = self.class_select_var.get().strip()
        if not class_name:
            messagebox.showwarning("No class", "Select or create a class first.")
            return

        try:
            class_id = int(class_name.split("#")[1]) if "#" in class_name else None
        except Exception:
            class_id = None
        if not class_id:
            messagebox.showerror("Error", "Invalid class selection.")
            return

        class_date_str = self.class_date_var.get().strip() or datetime.now().strftime("%Y-%m-%d")
        start_time = self.class_time_var.get().strip() or datetime.now().strftime("%I:%M %p")

        # Create a new session for this class
        session_number = self.db.create_class_session(class_id, class_date_str, start_time)

        self.current_class_id = class_id
        self.current_class_name = class_name.split("#")[0].strip()
        self.current_class_date = class_date_str
        self.current_session_number = session_number

        try:
            class_start_dt, late_cutoff = self._compute_class_times(class_date_str, start_time)
        except ValueError as exc:
            messagebox.showerror("Time", str(exc))
            return

        self.late_cutoff = late_cutoff
        self.scheduled_class_start = class_start_dt

        self._log(
            f"Class '{self.current_class_name}' Session #{session_number} scheduled for {class_start_dt.strftime('%H:%M')}. Waiting for teacher..."
        )

        # Log detector configuration (helpful to verify GPU/CUDA setup)
        try:
            import dlib  # type: ignore

            cuda_flag = bool(getattr(dlib, "DLIB_USE_CUDA", False))
            try:
                cuda_devices = int(dlib.cuda.get_num_devices()) if cuda_flag else 0
            except Exception:
                cuda_devices = 0
            self._log(
                f"Detector: model={ATTENDANCE_FACE_DETECTION_MODEL}, upsample={ATTENDANCE_FACE_DETECTION_UPSAMPLE}, "
                f"cnn_scale={ATTENDANCE_CNN_FRAME_SCALE} | dlib_cuda={cuda_flag}, cuda_devices={cuda_devices}"
            )
        except Exception:
            self._log(
                f"Detector: model={ATTENDANCE_FACE_DETECTION_MODEL}, upsample={ATTENDANCE_FACE_DETECTION_UPSAMPLE}, "
                f"cnn_scale={ATTENDANCE_CNN_FRAME_SCALE} | dlib_cuda=unknown"
            )
        
        roster = self.db.list_class_students(class_id)
        if not roster:
            messagebox.showwarning("No students", "This class has no students assigned.")
            return
        
        self._populate_roster(roster)
        self.stop_attendance_event = threading.Event()
        
        # Start timer thread to check for teacher arrival
        teacher_wait_thread = threading.Thread(
            target=self._wait_for_teacher, 
            args=(class_id, session_number, class_start_dt, roster, self.stop_attendance_event), 
            daemon=True
        )
        teacher_wait_thread.start()
    
    def _wait_for_teacher(
        self,
        class_id: int,
        session_number: int,
        scheduled_start: datetime,
        roster: List[Tuple[int, str]],
        stop_event: threading.Event,
    ) -> None:
        """Wait for teacher to arrive, cancel class if 20 minutes late"""
        try:
            max_wait_time = scheduled_start + timedelta(minutes=20)
            check_interval = 1

            teacher_arrived_event = threading.Event()
            self.teacher_arrived_event = teacher_arrived_event

            # Start camera in a separate thread
            camera_thread = threading.Thread(
                target=self._run_teacher_detection,
                args=(class_id, session_number, roster, teacher_arrived_event, stop_event),
                daemon=True,
            )
            camera_thread.start()

            # Wait for teacher or timeout
            while datetime.now() < max_wait_time and not stop_event.is_set():
                if teacher_arrived_event.is_set():
                    self._log("✓ Teacher detected, proceeding with class")
                    return
                time.sleep(check_interval)

            if not teacher_arrived_event.is_set() and not stop_event.is_set():
                stop_event.set()
                # Teacher never arrived: do not count this as a lecture/session.
                try:
                    self.db.delete_class_session(class_id, session_number)
                except Exception as del_exc:
                    self._log_exception("Failed to delete cancelled session", del_exc)
                self._log("❌ Class cancelled: Teacher didn't arrive within 20 minutes of scheduled time")
                self._show_message_async(
                    "warning",
                    "Class Cancelled",
                    "Class cancelled because teacher didn't arrive within 20 minutes.\n\n"
                    f"Scheduled: {scheduled_start.strftime('%H:%M')}\n"
                    f"Cancelled at: {datetime.now().strftime('%H:%M')}",
                )
        except Exception as exc:
            stop_event.set()
            self._log_exception("Teacher wait thread failed", exc)
        
    def _run_teacher_detection(
        self,
        class_id: int,
        session_number: int,
        roster: List[Tuple[int, str]],
        teacher_arrived_event: threading.Event,
        stop_event: threading.Event,
    ) -> None:
        """Modified camera session that waits for teacher first"""
        name_by_id = {sid: name for sid, name in roster}
        
        # Get student encodings
        known_encodings: List[np.ndarray] = []
        known_ids: List[int] = []
        known_types: List[str] = []
        for sid, _ in roster:
            encs = self.db.get_student_encodings(sid)
            known_encodings.extend(encs)
            known_ids.extend([sid] * len(encs))
            known_types.extend(["student"] * len(encs))
        
        # Get teacher encodings
        teacher_encodings, teacher_ids = self.db.get_all_teacher_encodings()
        known_encodings.extend(teacher_encodings)
        known_ids.extend(teacher_ids)
        known_types.extend(["teacher"] * len(teacher_encodings))
        known_encoding_matrix = np.array(known_encodings) if known_encodings else np.empty((0, 128))
        teacher_name_by_id = {tid: tname for tid, tname in self.db.list_teachers()}
        
        cap = cv2.VideoCapture(self._get_camera_index())
        if not cap.isOpened():
            self._log("❌ Cannot access camera")
            stop_event.set()
            return

        # Lower capture workload and reduce stale frame buffering.
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, ATTENDANCE_CAMERA_WIDTH)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, ATTENDANCE_CAMERA_HEIGHT)
        try:
            cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        except Exception:
            pass
        
        recognized_ids: Dict[int, datetime] = {}
        teacher_arrived = False
        teacher_arrival_time = None
        class_actual_start_time = None
        class_end_time = None
        unknown_faces_saved = set()
        unknown_zone_last_saved: Dict[str, float] = {}
        last_draw_items: List[Tuple[Tuple[int, int, int, int], str, Tuple[int, int, int]]] = []
        frame_counter = 0
        
        # Check if end time is set
        if hasattr(self, 'class_end_time_var'):
            end_time_str = self.class_end_time_var.get().strip()
            if end_time_str:
                try:
                    class_date_str = self.class_date_var.get().strip() or datetime.now().strftime("%Y-%m-%d")
                    class_end_time = datetime.strptime(f"{class_date_str} {end_time_str}", "%Y-%m-%d %I:%M %p")
                    self._log(f"⏰ Class will automatically end at {class_end_time.strftime('%H:%M')}")
                except:
                    self._log("⚠️  Invalid end time format, ignoring...")
        
        # Create directory for unknown visitors
        unknown_dir = Path("unknown_visitors")
        unknown_dir.mkdir(exist_ok=True)
        
        self._log("📷 Camera active. Waiting for teacher... Press C or ESC to stop.")

        # Resolve effective detector model + parameters (auto-fallback if CUDA isn't available)
        det_model = (ATTENDANCE_FACE_DETECTION_MODEL or "hog").strip().lower()
        det_upsample = int(ATTENDANCE_FACE_DETECTION_UPSAMPLE)
        det_scale = float(ATTENDANCE_CNN_FRAME_SCALE) if det_model == "cnn" else float(ATTENDANCE_FRAME_SCALE)
        if det_model == "cnn":
            try:
                import dlib  # type: ignore

                cuda_ok = bool(getattr(dlib, "DLIB_USE_CUDA", False))
                if cuda_ok:
                    try:
                        cuda_ok = int(dlib.cuda.get_num_devices()) > 0
                    except Exception:
                        cuda_ok = False
                if not cuda_ok:
                    self._log("⚠️  CNN requested but CUDA dlib not detected; falling back to HOG.")
                    det_model = "hog"
                    det_scale = float(ATTENDANCE_FRAME_SCALE)
            except Exception:
                self._log("⚠️  CNN requested but dlib CUDA check failed; falling back to HOG.")
                det_model = "hog"
                det_scale = float(ATTENDANCE_FRAME_SCALE)
        
        try:
            while True:
                if stop_event.is_set():
                    self._log("Attendance stop signal received. Closing camera session.")
                    break

                # Check if class should end
                if class_end_time and datetime.now() >= class_end_time:
                    self._log(f"⏰ Class time ended at {class_end_time.strftime('%H:%M')} - Auto-closing")
                    stop_event.set()
                    self._show_message_async("info", "Class Ended", f"Class automatically ended at {class_end_time.strftime('%H:%M')}")
                    break
                
                ret, frame = cap.read()
                if not ret:
                    self._log("⚠️  Camera frame read failed. Ending attendance loop.")
                    break

                frame_counter += 1
                if frame_counter % ATTENDANCE_PROCESS_EVERY_N_FRAMES == 0:
                    last_draw_items = []
                    small_frame = cv2.resize(frame, (0, 0), fx=det_scale, fy=det_scale)
                    rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                    try:
                        locations_small = face_recognition.face_locations(
                            rgb_small,
                            number_of_times_to_upsample=det_upsample,
                            model=det_model,
                        )
                    except Exception as det_exc:
                        if det_model == "cnn":
                            self._log(f"⚠️  CNN face detection failed ({det_exc}); falling back to HOG.")
                            det_model = "hog"
                            det_scale = float(ATTENDANCE_FRAME_SCALE)
                            small_frame = cv2.resize(frame, (0, 0), fx=det_scale, fy=det_scale)
                            rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                            locations_small = face_recognition.face_locations(
                                rgb_small,
                                number_of_times_to_upsample=det_upsample,
                                model=det_model,
                            )
                        else:
                            raise
                    encs = face_recognition.face_encodings(rgb_small, locations_small)
                    locations_small = locations_small[:MAX_FACES_PER_FRAME]
                    encs = encs[:MAX_FACES_PER_FRAME]

                    for enc, loc_small in zip(encs, locations_small):
                        try:
                            top_s, right_s, bottom_s, left_s = loc_small
                            top = int(top_s / det_scale)
                            right = int(right_s / det_scale)
                            bottom = int(bottom_s / det_scale)
                            left = int(left_s / det_scale)

                            top = max(0, min(frame.shape[0] - 1, top))
                            bottom = max(0, min(frame.shape[0], bottom))
                            left = max(0, min(frame.shape[1] - 1, left))
                            right = max(0, min(frame.shape[1], right))
                            if right <= left or bottom <= top:
                                continue

                            label = "Unknown"
                            box_color = (0, 0, 255)

                            matched = False
                            if known_encoding_matrix.size > 0:
                                distances = face_recognition.face_distance(known_encoding_matrix, enc)
                                if len(distances) > 0:
                                    best_idx = int(np.argmin(distances))
                                    if distances[best_idx] <= ENCODING_TOLERANCE:
                                        matched = True
                                        person_id = known_ids[best_idx]
                                        person_type = known_types[best_idx]

                                        if person_type == "teacher" and not teacher_arrived:
                                            teacher_arrived = True
                                            teacher_arrival_time = datetime.now()
                                            teacher_name = teacher_name_by_id.get(person_id, "Unknown Teacher")
                                            self.db.mark_teacher_attendance(
                                                class_id,
                                                person_id,
                                                teacher_arrival_time,
                                                session_number=session_number,
                                            )
                                            self._log(
                                                f"🎓 Teacher {teacher_name} arrived at {teacher_arrival_time.strftime('%H:%M:%S')}"
                                            )
                                            teacher_arrived_event.set()
                                            class_actual_start_time = teacher_arrival_time + timedelta(minutes=5)
                                            self._log(
                                                f"✓ Class will start at {class_actual_start_time.strftime('%H:%M:%S')} (5 min after teacher)"
                                            )
                                            self.late_cutoff = class_actual_start_time

                                        elif person_type == "student":
                                            sid = person_id
                                            if sid not in recognized_ids:
                                                recognized_ids[sid] = datetime.now()
                                                self._mark_roster_present(sid, recognized_ids[sid])

                                        label = "TEACHER" if person_type == "teacher" else name_by_id.get(person_id, "Student")
                                        box_color = (0, 255, 0)

                            if not matched:
                                zone_key = f"{(left + right) // 80}_{(top + bottom) // 80}"
                                now_epoch = time.time()
                                last_saved_epoch = unknown_zone_last_saved.get(zone_key, 0.0)

                                if now_epoch - last_saved_epoch >= UNKNOWN_VISITOR_SAVE_COOLDOWN_SEC:
                                    unknown_zone_last_saved[zone_key] = now_epoch
                                    now = datetime.now()
                                    visitor_id = self.db.save_unknown_visitor(
                                        class_id,
                                        "",
                                        now,
                                        enc,
                                        session_number=session_number,
                                    )
                                    visitor_key = f"visitor_{visitor_id}"

                                    row = self.db.conn.execute(
                                        "SELECT visit_count FROM unknown_visitors WHERE id=?",
                                        (visitor_id,),
                                    ).fetchone()
                                    visit_count = row[0] if row else 1

                                    if visitor_key not in unknown_faces_saved:
                                        if visit_count == 1:
                                            filename = f"unknown_{class_id}_{now.strftime('%Y%m%d_%H%M%S_%f')}.jpg"
                                            filepath = unknown_dir / filename

                                            face_img = frame[
                                                max(0, top - 20):min(frame.shape[0], bottom + 20),
                                                max(0, left - 20):min(frame.shape[1], right + 20),
                                            ]
                                            if face_img.size > 0:
                                                h, w = face_img.shape[:2]
                                                cv2.rectangle(face_img, (0, h - 25), (w, h), (0, 0, 0), -1)
                                                cv2.putText(
                                                    face_img,
                                                    now.strftime('%Y-%m-%d %H:%M:%S'),
                                                    (5, h - 8),
                                                    cv2.FONT_HERSHEY_SIMPLEX,
                                                    0.4,
                                                    (255, 255, 255),
                                                    1,
                                                )

                                                cv2.imwrite(str(filepath), face_img)
                                                self.db.conn.execute(
                                                    "UPDATE unknown_visitors SET image_path=? WHERE id=?",
                                                    (str(filepath), visitor_id),
                                                )
                                                self.db.conn.commit()
                                                self._log(f"⚠️  New unknown visitor #{visitor_id} - Saved: {filename}")
                                        else:
                                            self._log(f"⚠️  Unknown visitor #{visitor_id} returned (visit #{visit_count})")
                                        unknown_faces_saved.add(visitor_key)

                                    label = f"Unknown #{visitor_id} (Visit #{visit_count})"

                            last_draw_items.append(((top, right, bottom, left), label, box_color))
                        except Exception as face_exc:
                            self._log_exception("Error while processing detected face", face_exc)
                            continue

                for (top, right, bottom, left), label, box_color in last_draw_items:
                    cv2.rectangle(frame, (left, top), (right, bottom), box_color, 2)
                    cv2.putText(frame, label, (left, max(20, top - 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, box_color, 2)
                
                # Display status
                h, w = frame.shape[:2]
                cv2.rectangle(frame, (0, 0), (w, 120), (0, 0, 0), -1)
                
                if teacher_arrived:
                    cv2.putText(frame, f"Teacher arrived: {teacher_arrival_time.strftime('%H:%M:%S')}", 
                              (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                    if class_actual_start_time:
                        now = datetime.now()
                        if now < class_actual_start_time:
                            seconds_left = int((class_actual_start_time - now).total_seconds())
                            cv2.putText(frame, f"Class starts in: {seconds_left}s", 
                                      (10, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
                        else:
                            cv2.putText(frame, "Class in session", 
                                      (10, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                else:
                    cv2.putText(frame, "Waiting for teacher...", 
                              (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 165, 255), 2)
                    if hasattr(self, 'scheduled_class_start'):
                        cv2.putText(frame, f"Scheduled: {self.scheduled_class_start.strftime('%H:%M')}", 
                                  (10, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
                
                cv2.putText(frame, f"Students: {len(recognized_ids)}/{len(roster)}", 
                          (10, 100), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
                
                cv2.imshow("Attendance - Face Recognition", frame)
                key = cv2.waitKey(1) & 0xFF
                if key in (ord('c'), ord('C'), 27):
                    stop_event.set()
                    break
        except Exception as exc:
            stop_event.set()
            self._log_exception("Attendance camera loop crashed", exc)
            self._show_message_async(
                "error",
                "Attendance Error",
                "Camera loop crashed. Attendance session closed safely. Check attendance_debug.log",
            )
        finally:
            cap.release()
            cv2.destroyAllWindows()
            
            try:
                # Mark attendance for recognized students
                if recognized_ids:
                    session_num = getattr(self, 'current_session_number', None)
                    for sid, ts in recognized_ids.items():
                        self.db.mark_attendance(class_id, sid, ts, session_number)

                # Show final report
                absentees = self.db.absent_for_class(class_id)
                if absentees:
                    self._log(f"Absent: {', '.join(absentees)}")
                    self._mark_absent_roster(absentees)
                else:
                    self._log("✓ All students present")

                self._log("✓ Attendance session completed")
            except Exception as finalize_exc:
                self._log_exception("Failed to finalize attendance session", finalize_exc)

    def _run_camera_session(self, class_id: int, roster: List[Tuple[int, str]]) -> Dict[int, datetime]:
        name_by_id = {sid: name for sid, name in roster}
        
        # Get student encodings
        known_encodings: List[np.ndarray] = []
        known_ids: List[int] = []
        known_types: List[str] = []  # "student" or "teacher"
        for sid, _ in roster:
            encs = self.db.get_student_encodings(sid)
            known_encodings.extend(encs)
            known_ids.extend([sid] * len(encs))
            known_types.extend(["student"] * len(encs))
        
        # Get teacher encodings
        teacher_encodings, teacher_ids = self.db.get_all_teacher_encodings()
        known_encodings.extend(teacher_encodings)
        known_ids.extend(teacher_ids)
        known_types.extend(["teacher"] * len(teacher_encodings))
        
        cap = cv2.VideoCapture(self._get_camera_index())
        if not cap.isOpened():
            raise RuntimeError("Cannot access camera.")
        
        recognized_ids: Dict[int, datetime] = {}
        teacher_arrived = False
        teacher_arrival_time = None
        class_actual_start_time = None
        unknown_faces_saved = set()  # Track saved unknown faces by frame signature
        
        # Create directory for unknown visitors
        unknown_dir = Path("unknown_visitors")
        unknown_dir.mkdir(exist_ok=True)
        
        self._log("Press C or ESC to stop the camera window.")

        det_model = (ATTENDANCE_FACE_DETECTION_MODEL or "hog").strip().lower()
        det_upsample = int(ATTENDANCE_FACE_DETECTION_UPSAMPLE)
        det_scale = float(ATTENDANCE_CNN_FRAME_SCALE) if det_model == "cnn" else 1.0
        
        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    break

                if det_scale != 1.0:
                    small_frame = cv2.resize(frame, (0, 0), fx=det_scale, fy=det_scale)
                    rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                    try:
                        locations_small = face_recognition.face_locations(
                            rgb_small,
                            number_of_times_to_upsample=det_upsample,
                            model=det_model,
                        )
                    except Exception as det_exc:
                        if det_model == "cnn":
                            self._log(f"⚠️  CNN face detection failed ({det_exc}); falling back to HOG.")
                            det_model = "hog"
                            det_scale = 1.0
                            continue
                        else:
                            raise
                    encs = face_recognition.face_encodings(rgb_small, locations_small)

                    # Scale locations back up for drawing/cropping on the original frame.
                    inv = 1.0 / det_scale
                    locations = [
                        (
                            int(top * inv),
                            int(right * inv),
                            int(bottom * inv),
                            int(left * inv),
                        )
                        for (top, right, bottom, left) in locations_small
                    ]
                else:
                    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    try:
                        locations = face_recognition.face_locations(
                            rgb,
                            number_of_times_to_upsample=det_upsample,
                            model=det_model,
                        )
                    except Exception as det_exc:
                        if det_model == "cnn":
                            self._log(f"⚠️  CNN face detection failed ({det_exc}); falling back to HOG.")
                            det_model = "hog"
                            locations = face_recognition.face_locations(
                                rgb,
                                number_of_times_to_upsample=det_upsample,
                                model=det_model,
                            )
                        else:
                            raise
                    encs = face_recognition.face_encodings(rgb, locations)

                locations = locations[:MAX_FACES_PER_FRAME]
                encs = encs[:MAX_FACES_PER_FRAME]
                
                for enc, loc in zip(encs, locations):
                    distances = face_recognition.face_distance(known_encodings, enc)
                    if len(distances) == 0:
                        continue
                    
                    best_idx = int(np.argmin(distances))
                    if distances[best_idx] <= ENCODING_TOLERANCE:
                        person_id = known_ids[best_idx]
                        person_type = known_types[best_idx]
                        
                        if person_type == "teacher" and not teacher_arrived:
                            # Teacher arrived!
                            teacher_arrived = True
                            teacher_arrival_time = datetime.now()
                            teachers = self.db.list_teachers()
                            teacher_name = next((name for tid, name in teachers if tid == person_id), "Unknown Teacher")
                            self.db.mark_teacher_attendance(class_id, person_id, teacher_arrival_time)
                            self._log(f"🎓 Teacher {teacher_name} arrived at {teacher_arrival_time.strftime('%H:%M:%S')}")
                            
                            # Class will start 5 minutes after teacher arrival
                            class_actual_start_time = teacher_arrival_time + timedelta(minutes=5)
                            self._log(f"✓ Class will start at {class_actual_start_time.strftime('%H:%M:%S')} (5 min after teacher)")
                            
                            # Update late cutoff based on actual class start time
                            self.late_cutoff = class_actual_start_time + timedelta(minutes=1)
                        
                        elif person_type == "student":
                            # Mark students even if they arrive before the teacher; they'll be on time if before start
                            sid = person_id
                            if sid not in recognized_ids:
                                recognized_ids[sid] = datetime.now()
                                self._mark_roster_present(sid, recognized_ids[sid])
                        
                        top, right, bottom, left = loc
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                        if person_type == "teacher":
                            label = f"TEACHER - ID {person_id}"
                        else:
                            label = f"ID {person_id} - {name_by_id.get(person_id, '')}"
                        cv2.putText(frame, label, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                    else:
                        # Unknown person detected
                        top, right, bottom, left = loc
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                        cv2.putText(frame, "Unknown Visitor", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                        
                        # Create a signature for this face to avoid duplicate saves
                        face_signature = f"{left}_{top}_{right}_{bottom}_{int(time.time())}"
                        if face_signature not in unknown_faces_saved:
                            # Save the unknown face image
                            now = datetime.now()
                            filename = f"unknown_{class_id}_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
                            filepath = unknown_dir / filename
                            
                            # Extract face region and save
                            face_img = frame[top:bottom, left:right]
                            # Add timestamp overlay
                            face_with_timestamp = face_img.copy()
                            h, w = face_with_timestamp.shape[:2]
                            cv2.rectangle(face_with_timestamp, (0, h-30), (w, h), (0, 0, 0), -1)
                            cv2.putText(face_with_timestamp, now.strftime('%Y-%m-%d %H:%M:%S'), 
                                      (5, h-10), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 1)
                            
                            cv2.imwrite(str(filepath), face_with_timestamp)
                            self.db.save_unknown_visitor(class_id, str(filepath), now)
                            self._log(f"⚠️  Unknown visitor detected at {now.strftime('%H:%M:%S')} - Image saved: {filename}")
                            unknown_faces_saved.add(face_signature)
                
                # Display status on frame
                h, w = frame.shape[:2]
                cv2.rectangle(frame, (0, 0), (w, 100), (0, 0, 0), -1)
                if teacher_arrived:
                    cv2.putText(frame, f"Teacher arrived: {teacher_arrival_time.strftime('%H:%M:%S')}", 
                              (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                    if class_actual_start_time:
                        cv2.putText(frame, f"Class starts: {class_actual_start_time.strftime('%H:%M:%S')}", 
                                  (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                else:
                    cv2.putText(frame, "Waiting for teacher...", 
                              (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 2)
                
                cv2.putText(frame, f"Students present: {len(recognized_ids)}/{len(roster)}", 
                          (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
                
                cv2.imshow("Attendance", frame)
                key = cv2.waitKey(1) & 0xFF
                if key in (ord('c'), 27):
                    break
        finally:
            cap.release()
            cv2.destroyAllWindows()
        
        return recognized_ids

    def _run_on_ui_thread(self, callback, *args) -> None:
        if threading.current_thread() is threading.main_thread():
            callback(*args)
        else:
            self.root.after(0, lambda: callback(*args))

    def _show_message_async(self, level: str, title: str, message: str) -> None:
        handlers = {
            "info": messagebox.showinfo,
            "warning": messagebox.showwarning,
            "error": messagebox.showerror,
        }
        show_fn = handlers.get(level, messagebox.showinfo)
        self._run_on_ui_thread(show_fn, title, message)

    def _append_log_to_ui(self, msg: str) -> None:
        if not hasattr(self, "log_text"):
            return
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{ts}] {msg}\n")
        self.log_text.configure(state="disabled")
        self.log_text.see("end")

    def _log_exception(self, context: str, exc: Exception) -> None:
        LOGGER.exception("%s: %s", context, exc)
        self._log(f"❌ {context}: {exc}")

    def _log(self, msg: str) -> None:
        LOGGER.info(msg)
        self._run_on_ui_thread(self._append_log_to_ui, msg)

    def _compute_class_times(self, date_str: str, time_str: str) -> Tuple[datetime, datetime]:
        try:
            start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %I:%M %p")
        except ValueError as exc:
            raise ValueError("Invalid date/time format. Use YYYY-MM-DD and hh:mm AM/PM.") from exc
        late_cutoff = start_dt + timedelta(minutes=15)
        return start_dt, late_cutoff

    def _export_csv(self) -> None:
        if not getattr(self, "current_class_id", None):
            messagebox.showinfo("Export", "Run an attendance session first.")
            return
        class_id = self.current_class_id
        class_name = getattr(self, "current_class_name", "class")
        class_date = getattr(self, "current_class_date", datetime.now().strftime("%Y-%m-%d"))
        roster = self.db.list_class_students(class_id)
        name_by_id = {sid: name for sid, name in roster}
        session_number = getattr(self, "current_session_number", None)
        if session_number:
            attendance_rows = self.db.conn.execute(
                "SELECT s.id, s.name, a.timestamp FROM attendance a JOIN students s ON s.id=a.student_id "
                "WHERE a.class_id=? AND a.session_number=? ORDER BY a.timestamp",
                (class_id, int(session_number)),
            ).fetchall()
        else:
            attendance_rows = self.db.conn.execute(
                "SELECT s.id, s.name, a.timestamp FROM attendance a JOIN students s ON s.id=a.student_id "
                "WHERE a.class_id=? ORDER BY a.timestamp",
                (class_id,),
            ).fetchall()

        arrival_by_id = {int(sid): ts for sid, _name, ts in attendance_rows}
        late_cutoff = getattr(self, "late_cutoff", None)

        warning_rows = self.db.get_warning_summary_for_class(class_id)
        warning_by_id = {sid: (total, attended, absent, warning) for sid, _n, total, attended, absent, warning in warning_rows}

        rows = []
        for sid, name in roster:
            ts_str = arrival_by_id.get(int(sid))
            total_lect, attended_lect, absent_lect, warning = warning_by_id.get(int(sid), (0, 0, 0, ""))
            if ts_str:
                ts = datetime.fromisoformat(ts_str)
                status = "Present"
                if late_cutoff and ts > late_cutoff:
                    status = "Late"
                rows.append([name, status, ts.strftime("%Y-%m-%d %H:%M:%S"), class_name, class_date, total_lect, absent_lect, warning])
            else:
                rows.append([name, "Absent", "", class_name, class_date, total_lect, absent_lect, warning])

        default_name = f"attendance_{class_name.replace(' ', '_')}_{class_date}.csv"
        filename = filedialog.asksaveasfilename(
            title="Save Attendance CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not filename:
            return

        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Student", "Status", "Arrival", "Class", "Date", "Total Lectures", "Absent Lectures", "Warning"])
            writer.writerows(rows)
        self._log(f"Exported CSV to {filename}")
        messagebox.showinfo("Export", f"Saved {filename}")

    def _export_excel(self) -> None:
        if not HAS_OPENPYXL:
            messagebox.showerror("Excel", "openpyxl not installed. Run: pip install openpyxl")
            return
        if not getattr(self, "current_class_id", None):
            messagebox.showinfo("Export", "Run an attendance session first.")
            return
        class_id = self.current_class_id
        class_name = getattr(self, "current_class_name", "class")
        class_date = getattr(self, "current_class_date", datetime.now().strftime("%Y-%m-%d"))
        roster = self.db.list_class_students(class_id)
        session_number = getattr(self, "current_session_number", None)
        if session_number:
            attendance_rows = self.db.conn.execute(
                "SELECT s.id, s.name, a.timestamp FROM attendance a JOIN students s ON s.id=a.student_id "
                "WHERE a.class_id=? AND a.session_number=? ORDER BY a.timestamp",
                (class_id, int(session_number)),
            ).fetchall()
        else:
            attendance_rows = self.db.conn.execute(
                "SELECT s.id, s.name, a.timestamp FROM attendance a JOIN students s ON s.id=a.student_id "
                "WHERE a.class_id=? ORDER BY a.timestamp",
                (class_id,),
            ).fetchall()

        arrival_by_id = {int(sid): ts for sid, _name, ts in attendance_rows}
        late_cutoff = getattr(self, "late_cutoff", None)

        warning_rows = self.db.get_warning_summary_for_class(class_id)
        warning_by_id = {sid: (total, attended, absent, warning) for sid, _n, total, attended, absent, warning in warning_rows}

        # Get teacher attendance
        if session_number:
            teacher_attendance = self.db.conn.execute(
                "SELECT t.name, ta.timestamp FROM teacher_attendance ta "
                "JOIN teachers t ON ta.teacher_id = t.id "
                "WHERE ta.class_id = ? AND ta.session_number = ? ORDER BY ta.timestamp",
                (class_id, int(session_number)),
            ).fetchall()
        else:
            teacher_attendance = self.db.conn.execute(
                "SELECT t.name, ta.timestamp FROM teacher_attendance ta "
                "JOIN teachers t ON ta.teacher_id = t.id "
                "WHERE ta.class_id = ? ORDER BY ta.timestamp",
                (class_id,),
            ).fetchall()

        # Get unknown visitors
        if session_number:
            unknown_visitors = self.db.conn.execute(
                "SELECT timestamp, image_path FROM unknown_visitors "
                "WHERE class_id = ? AND session_number = ? ORDER BY timestamp",
                (class_id, int(session_number)),
            ).fetchall()
        else:
            unknown_visitors = self.db.conn.execute(
                "SELECT timestamp, image_path FROM unknown_visitors "
                "WHERE class_id = ? ORDER BY timestamp",
                (class_id,),
            ).fetchall()

        wb = Workbook()
        
        # Sheet 1: Student Attendance
        ws = wb.active
        ws.title = "Students"
        ws.append(["Student", "Status", "Arrival Time", "Class", "Date", "Total Lectures", "Absent Lectures", "Warning"])
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for sid, name in roster:
            total_lect, _attended_lect, absent_lect, warning = warning_by_id.get(int(sid), (0, 0, 0, ""))
            ts_str = arrival_by_id.get(int(sid))
            if ts_str:
                ts = datetime.fromisoformat(ts_str)
                status = "Present"
                color = "C6EFCE"
                if late_cutoff and ts > late_cutoff:
                    status = "Late"
                    color = "FFC7CE"
                row_num = ws.max_row + 1
                ws.append([name, status, ts.strftime("%Y-%m-%d %H:%M:%S"), class_name, class_date, total_lect, absent_lect, warning])
                ws[f"B{row_num}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            else:
                row_num = ws.max_row + 1
                ws.append([name, "Absent", "", class_name, class_date, total_lect, absent_lect, warning])
                ws[f"B{row_num}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f"B{row_num}"].font = Font(color="FFFFFF")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 16

        # Sheet 1b: Warnings summary (per class)
        ws_warn = wb.create_sheet("Warnings")
        ws_warn.append(["Student", "Total Lectures", "Attended Lectures", "Absent Lectures", "Warning", "Class"])
        for cell in ws_warn[1]:
            cell.fill = header_fill
            cell.font = header_font
        for _sid, student_name, total_lect, attended_lect, absent_lect, warning in warning_rows:
            ws_warn.append([student_name, total_lect, attended_lect, absent_lect, warning, class_name])
        ws_warn.column_dimensions["A"].width = 25
        ws_warn.column_dimensions["B"].width = 14
        ws_warn.column_dimensions["C"].width = 16
        ws_warn.column_dimensions["D"].width = 14
        ws_warn.column_dimensions["E"].width = 14
        ws_warn.column_dimensions["F"].width = 20

        # Sheet 2: Teacher Attendance
        ws_teacher = wb.create_sheet("Teachers")
        ws_teacher.append(["Teacher Name", "Arrival Time", "Class", "Date"])
        for cell in ws_teacher[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        if teacher_attendance:
            for teacher_name, ts_str in teacher_attendance:
                ts = datetime.fromisoformat(ts_str)
                ws_teacher.append([teacher_name, ts.strftime("%Y-%m-%d %H:%M:%S"), class_name, class_date])
        else:
            ws_teacher.append(["No teacher recorded", "", class_name, class_date])
        
        ws_teacher.column_dimensions["A"].width = 20
        ws_teacher.column_dimensions["B"].width = 20
        ws_teacher.column_dimensions["C"].width = 15
        ws_teacher.column_dimensions["D"].width = 12

        # Sheet 3: Unknown Visitors
        ws_visitors = wb.create_sheet("Unknown Visitors")
        ws_visitors.append(["Visitor #", "Arrival Time", "Image Path", "Class", "Date"])
        for cell in ws_visitors[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        if unknown_visitors:
            for idx, (ts_str, image_path) in enumerate(unknown_visitors, 1):
                ts = datetime.fromisoformat(ts_str)
                ws_visitors.append([f"Visitor {idx}", ts.strftime("%Y-%m-%d %H:%M:%S"), 
                                   image_path, class_name, class_date])
                # Make the row orange to highlight unknown visitors
                row_num = ws_visitors.max_row
                for col in ["A", "B", "C", "D", "E"]:
                    ws_visitors[f"{col}{row_num}"].fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
        else:
            ws_visitors.append(["No unknown visitors", "", "", class_name, class_date])
        
        ws_visitors.column_dimensions["A"].width = 12
        ws_visitors.column_dimensions["B"].width = 20
        ws_visitors.column_dimensions["C"].width = 35
        ws_visitors.column_dimensions["D"].width = 15
        ws_visitors.column_dimensions["E"].width = 12

        default_name = f"attendance_{class_name.replace(' ', '_')}_{class_date}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Save Attendance Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if not filename:
            return

        wb.save(filename)
        self._log(f"Exported Excel to {filename}")
        messagebox.showinfo("Export", f"Saved {filename}\n\nشامل:\n- قوتابیان\n- مامۆستا\n- میوانە نەناسراوەکان")

    def _export_csv_detailed(self) -> None:
        """Export detailed attendance with teachers and unknown visitors"""
        if not getattr(self, "current_class_id", None):
            messagebox.showinfo("Export", "Run an attendance session first.")
            return
        
        class_id = self.current_class_id
        class_name = getattr(self, "current_class_name", "class")
        class_date = getattr(self, "current_class_date", datetime.now().strftime("%Y-%m-%d"))
        
        # Get all data
        roster = self.db.list_class_students(class_id)
        attendance = self.db.attendance_for_class(class_id)
        arrival_by_name = {name: ts for name, ts in attendance}
        late_cutoff = getattr(self, "late_cutoff", None)
        
        teacher_attendance = self.db.conn.execute(
            "SELECT t.name, ta.timestamp FROM teacher_attendance ta "
            "JOIN teachers t ON ta.teacher_id = t.id "
            "WHERE ta.class_id = ? ORDER BY ta.timestamp",
            (class_id,)
        ).fetchall()
        
        unknown_visitors = self.db.conn.execute(
            "SELECT timestamp, image_path FROM unknown_visitors "
            "WHERE class_id = ? ORDER BY timestamp",
            (class_id,)
        ).fetchall()
        
        default_name = f"attendance_detailed_{class_name.replace(' ', '_')}_{class_date}.csv"
        filename = filedialog.asksaveasfilename(
            title="Save Detailed Report CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not filename:
            return
        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            
            # Students section
            writer.writerow(["=== STUDENT ATTENDANCE ==="])
            writer.writerow(["Student", "Status", "Arrival", "Class", "Date"])
            for sid, name in roster:
                ts_str = arrival_by_name.get(name)
                if ts_str:
                    ts = datetime.fromisoformat(ts_str)
                    status = "Present"
                    if late_cutoff and ts > late_cutoff:
                        status = "Late"
                    writer.writerow([name, status, ts.strftime("%Y-%m-%d %H:%M:%S"), class_name, class_date])
                else:
                    writer.writerow([name, "Absent", "", class_name, class_date])
            
            # Teachers section
            writer.writerow([])
            writer.writerow(["=== TEACHER ATTENDANCE ==="])
            writer.writerow(["Teacher Name", "Arrival Time", "Class", "Date"])
            if teacher_attendance:
                for teacher_name, ts_str in teacher_attendance:
                    ts = datetime.fromisoformat(ts_str)
                    writer.writerow([teacher_name, ts.strftime("%Y-%m-%d %H:%M:%S"), class_name, class_date])
            else:
                writer.writerow(["No teacher recorded", "", class_name, class_date])
            
            # Unknown visitors section
            writer.writerow([])
            writer.writerow(["=== UNKNOWN VISITORS ==="])
            writer.writerow(["Visitor #", "Arrival Time", "Image Path", "Class", "Date"])
            if unknown_visitors:
                for idx, (ts_str, image_path) in enumerate(unknown_visitors, 1):
                    ts = datetime.fromisoformat(ts_str)
                    writer.writerow([f"Visitor {idx}", ts.strftime("%Y-%m-%d %H:%M:%S"), 
                                   image_path, class_name, class_date])
            else:
                writer.writerow(["No unknown visitors", "", "", class_name, class_date])
        
        self._log(f"Exported detailed CSV to {filename}")
        messagebox.showinfo("Export", f"Saved {filename}")

    def _edit_student_dialog(self) -> None:
        students = self.db.list_students()
        if not students:
            messagebox.showinfo("Info", "No students to edit.")
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Student")
        dialog.geometry("400x350")
        tk.Label(dialog, text="Select Student:").pack(pady=4)
        listbox = tk.Listbox(dialog, height=8)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for sid, sname in students:
            listbox.insert(tk.END, f"{sname} #{sid}")
        tk.Label(dialog, text="New Name:").pack(pady=4)
        name_var = tk.StringVar()
        tk.Entry(dialog, textvariable=name_var, width=30).pack(pady=4)
        tk.Label(dialog, text="Select Classes:").pack(pady=4)
        class_listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, height=5)
        class_listbox.pack(fill="both", expand=True, padx=10, pady=4)
        classes = self.db.list_classes()
        for class_id, class_name in classes:
            class_listbox.insert(tk.END, f"{class_name} #{class_id}")
        def update_student_info() -> None:
            sel = listbox.curselection()
            if not sel:
                messagebox.showerror("Error", "Select a student.")
                return
            student_id = int(listbox.get(sel[0]).split("#")[1])
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showerror("Error", "Enter a name.")
                return
            selected_classes = class_listbox.curselection()
            class_ids = [int(class_listbox.get(i).split("#")[1]) for i in selected_classes]
            self.db.update_student(student_id, new_name, class_ids)
            self._log(f"Updated student {new_name} with {len(class_ids)} class(es).")
            dialog.destroy()
        tk.Button(dialog, text="Update", command=update_student_info).pack(pady=4)

    def _delete_student_dialog(self) -> None:
        students = self.db.list_students()
        if not students:
            messagebox.showinfo("Info", "No students to delete.")
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Delete Student")
        dialog.geometry("300x250")
        tk.Label(dialog, text="Select Student:").pack(pady=4)
        listbox = tk.Listbox(dialog, height=10)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for sid, sname in students:
            listbox.insert(tk.END, f"{sname} #{sid}")
        def delete_student_confirm() -> None:
            sel = listbox.curselection()
            if not sel:
                messagebox.showerror("Error", "Select a student.")
                return
            student_name = listbox.get(sel[0]).split("#")[0]
            student_id = int(listbox.get(sel[0]).split("#")[1])
            if messagebox.askyesno("Delete", f"Delete student '{student_name}'?"):
                self.db.delete_student(student_id)
                self._log(f"Deleted student '{student_name}'.")
                dialog.destroy()
        tk.Button(dialog, text="Delete", command=delete_student_confirm).pack(pady=4)

    def _populate_roster(self, roster: List[Tuple[int, str]]) -> None:
        if not hasattr(self, "roster_table"):
            return
        for item in self.roster_table.get_children():
            self.roster_table.delete(item)
        self.roster_index.clear()
        for sid, name in roster:
            item_id = self.roster_table.insert("", tk.END, values=(name, ""), tags=("pending",))
            self.roster_index[sid] = item_id

    def _mark_roster_present(self, student_id: int, ts: datetime) -> None:
        if threading.current_thread() is not threading.main_thread():
            self._run_on_ui_thread(self._mark_roster_present, student_id, ts)
            return
        if student_id in self.roster_index:
            item_id = self.roster_index[student_id]
            tag = "present_on_time" if ts <= self.late_cutoff else "present_late"
            self.roster_table.item(item_id, values=(self.roster_table.set(item_id, "name"), ts.strftime("%Y-%m-%d %H:%M:%S")), tags=(tag,))

    def _mark_absent_roster(self, absent_names: List[str]) -> None:
        if threading.current_thread() is not threading.main_thread():
            self._run_on_ui_thread(self._mark_absent_roster, list(absent_names))
            return
        name_to_id = {self.roster_table.set(item_id, "name"): item_id for item_id in self.roster_table.get_children()}
        for name in absent_names:
            item_id = name_to_id.get(name)
            if item_id:
                self.roster_table.item(item_id, tags=("absent",))

    def _refresh_class_dropdown(self) -> None:
        classes = self.db.list_classes()
        menu = self.class_dropdown["menu"]
        menu.delete(0, tk.END)
        for class_id, class_name in classes:
            menu.add_command(label=f"{class_name} #{class_id}", command=tk._setit(self.class_select_var, f"{class_name} #{class_id}"))
        if classes:
            self.class_select_var.set(f"{classes[0][1]} #{classes[0][0]}")
        self._log(f"Loaded {len(classes)} class(es).")

    def _create_class_dialog(self) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title("Create New Class")
        dialog.geometry("350x200")
        dialog.resizable(False, False)
        
        pad = {"padx": 10, "pady": 8}
        
        # Class Name
        tk.Label(dialog, text="Class Name:", font=("Arial", 11, "bold")).pack(**pad, anchor="w")
        class_name_var = tk.StringVar()
        tk.Entry(dialog, textvariable=class_name_var, font=("Arial", 11), width=25).pack(fill="x", **pad)
        
        # Start Time
        tk.Label(dialog, text="Start Time (HH:MM):", font=("Arial", 11, "bold")).pack(**pad, anchor="w")
        time_var = tk.StringVar(value="10:00")
        tk.Entry(dialog, textvariable=time_var, font=("Arial", 11), width=25).pack(fill="x", **pad)
        
        # Buttons
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill="x", **pad)
        
        def save_class():
            name = class_name_var.get().strip()
            if not name:
                messagebox.showerror("Error", "Enter a class name")
                return
            # Create empty class, students can be added later
            class_id = self.db.create_class_definition(name, [])
            self._log(f"✓ Class '{name}' created successfully")
            self._refresh_class_list()
            self._refresh_class_dropdown()
            messagebox.showinfo("Success", f"Class '{name}' created!\n\nYou can add students to this class by editing it.")
            dialog.destroy()
        
        tk.Button(btn_frame, text="Create Class", command=save_class, width=15, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)

    def _edit_class_dialog(self) -> None:
        selection = self.class_listbox.curselection()
        if not selection:
            messagebox.showwarning("Error", "Select a class to edit")
            return
        
        # Use class_list_items mapping to handle session-based listbox
        if hasattr(self, 'class_list_items') and selection[0] < len(self.class_list_items):
            sel = self.class_list_items[selection[0]]
            class_id = sel['class_id']
            class_name = sel['class_name']
        else:
            classes = self.db.conn.execute("SELECT id, name FROM classes ORDER BY name").fetchall()
            if selection[0] >= len(classes):
                return
            class_id, class_name = classes[selection[0]]
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Class")
        dialog.geometry("620x520")
        dialog.minsize(620, 520)
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        
        pad = {"padx": 10, "pady": 8}
        
        # Class Name Section
        name_frame = tk.LabelFrame(dialog, text="Class Name", font=("Arial", 10, "bold"))
        name_frame.pack(fill="x", **pad)
        
        name_var = tk.StringVar(value=class_name)
        tk.Entry(name_frame, textvariable=name_var, font=("Arial", 11), width=40).pack(fill="x", **pad)
        
        # Students Selection
        students_frame_label = tk.LabelFrame(dialog, text="Assign Students", font=("Arial", 10, "bold"))
        students_frame_label.pack(fill="both", expand=True, **pad)
        
        tk.Label(students_frame_label, text="Select students (click to toggle):", font=("Arial", 9)).pack(padx=10, pady=5, anchor="w")
        
        # Scrollable canvas for students
        student_canvas = tk.Canvas(students_frame_label, height=150)
        student_scrollbar = tk.Scrollbar(students_frame_label, orient="vertical", command=student_canvas.yview)
        student_scrollable_frame = tk.Frame(student_canvas)
        
        student_scrollable_frame.bind(
            "<Configure>",
            lambda e: student_canvas.configure(scrollregion=student_canvas.bbox("all"))
        )
        
        student_canvas.create_window((0, 0), window=student_scrollable_frame, anchor="nw")
        student_canvas.configure(yscrollcommand=student_scrollbar.set)
        
        student_canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        student_scrollbar.pack(side="right", fill="y")
        
        # Add mousewheel scrolling
        def _on_mousewheel(event):
            student_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        student_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Load students
        all_students = self.db.list_students()
        current_student_ids = {sid for sid, _ in self.db.list_class_students(class_id)}
        
        student_vars = {}
        
        if all_students:
            for sid, sname in all_students:
                var = tk.BooleanVar(value=(sid in current_student_ids))
                tk.Checkbutton(student_scrollable_frame, text=f"  {sname}", variable=var, font=("Arial", 10), anchor="w").pack(fill="x", padx=10, pady=2)
                student_vars[sid] = var
        else:
            tk.Label(student_scrollable_frame, text="No students created yet. Create students first!", font=("Arial", 9), fg="red").pack(padx=10, pady=10)
        
        # Teachers Selection
        teachers_frame_label = tk.LabelFrame(dialog, text="Assign Teachers", font=("Arial", 10, "bold"))
        teachers_frame_label.pack(fill="both", expand=False, **pad)
        
        tk.Label(teachers_frame_label, text="Select teachers:", font=("Arial", 9)).pack(**pad, anchor="w")
        
        # Load teachers
        all_teachers = self.db.list_teachers()
        current_teacher_ids = {tid for tid, in self.db.conn.execute(
            "SELECT teacher_id FROM teacher_classes WHERE class_id = ?", (class_id,)
        ).fetchall()}
        
        teacher_vars = {}
        
        if all_teachers:
            for tid, tname in all_teachers:
                var = tk.BooleanVar(value=(tid in current_teacher_ids))
                tk.Checkbutton(teachers_frame_label, text=f"  {tname}", variable=var, font=("Arial", 10), anchor="w").pack(fill="x", **pad)
                teacher_vars[tid] = var
        else:
            tk.Label(teachers_frame_label, text="No teachers created yet. Create teachers first!", font=("Arial", 9), fg="red").pack(**pad)
        
        # Buttons
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill="x", pady=12, padx=14)
        
        def save_edit():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showerror("Error", "Class name cannot be empty")
                return
            
            # Get selected students and teachers
            selected_student_ids = [sid for sid, var in student_vars.items() if var.get()]
            selected_teacher_ids = [tid for tid, var in teacher_vars.items() if var.get()]
            
            # Update class name
            self.db.conn.execute("UPDATE classes SET name = ? WHERE id = ?", (new_name, class_id))
            
            # Update students
            self.db.update_class_students(class_id, selected_student_ids)
            
            # Update teachers for this class
            cur = self.db.conn.cursor()
            cur.execute("DELETE FROM teacher_classes WHERE class_id = ?", (class_id,))
            for tid in selected_teacher_ids:
                cur.execute("INSERT INTO teacher_classes (teacher_id, class_id) VALUES (?, ?)", (tid, class_id))
            
            self.db.conn.commit()
            
            self._log(f"✓ Class '{new_name}' updated: {len(selected_student_ids)} students, {len(selected_teacher_ids)} teachers")
            self._refresh_class_list()
            self._refresh_class_dropdown()
            messagebox.showinfo("Success", f"Class '{new_name}' updated!")
            dialog.destroy()
        
        tk.Button(btn_frame, text="Save Changes", command=save_edit, width=20, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=20).pack(side=tk.LEFT, padx=5)

    def _delete_class_dialog(self) -> None:
        selection = self.class_listbox.curselection()
        if not selection:
            messagebox.showwarning("Error", "Select a class to delete")
            return
        
        # Use class_list_items mapping to handle session-based listbox
        if hasattr(self, 'class_list_items') and selection[0] < len(self.class_list_items):
            sel = self.class_list_items[selection[0]]
            class_id = sel['class_id']
            class_name = sel['class_name']
        else:
            classes = self.db.conn.execute("SELECT id, name FROM classes ORDER BY name").fetchall()
            if selection[0] >= len(classes):
                return
            class_id, class_name = classes[selection[0]]
        
        if messagebox.askyesno("Delete", f"Delete class '{class_name}'?\n\nThis will remove all attendance records for this class."):
            self.db.delete_class(class_id)
            self._log(f"✓ Class '{class_name}' deleted")
            self._refresh_class_list()
            self._refresh_class_dropdown()
            messagebox.showinfo("Success", "Class deleted")

    def _get_camera_index(self) -> int:
        try:
            return int(self.camera_index_var.get())
        except Exception:
            messagebox.showwarning("Camera", "Invalid camera index. Using 0.")
            self.camera_index_var.set("0")
            return 0
    
    def _test_camera(self) -> None:
        """Test camera and show preview"""
        cam_index = self._get_camera_index()
        cap = cv2.VideoCapture(cam_index)
        
        if not cap.isOpened():
            messagebox.showerror("Camera Error", f"Camera {cam_index} is not available!\n\nTry different index:\n- 0 = Default camera\n- 1 = External/USB camera\n- 2 = Other camera\n\nUpdate the index and try again.")
            return
        
        self._log(f"✓ Camera {cam_index} is working! Press ESC to close the preview.")
        
        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                
                h, w = frame.shape[:2]
                cv2.rectangle(frame, (0, 0), (w, 80), (0, 0, 0), -1)
                cv2.putText(frame, f"Camera {cam_index} Preview - Press ESC to close", (15, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                cv2.putText(frame, "Camera is working! ✓", (15, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 1)
                
                cv2.imshow(f"Camera {cam_index} Test", frame)
                
                key = cv2.waitKey(1) & 0xFF
                if key == 27:  # ESC
                    break
        finally:
            cap.release()
            cv2.destroyAllWindows()
            self._log(f"Camera {cam_index} test closed.")


    def _refresh_contact_list(self, contact_listbox) -> None:
        """Refresh the contact list with all students"""
        contact_listbox.delete(0, tk.END)
        students = self.db.conn.execute("SELECT id, name FROM students ORDER BY name").fetchall()
        for student_id, name in students:
            contact_listbox.insert(tk.END, f"{name} (ID: {student_id})")
    
    def _edit_contact_info_dialog(self, contact_listbox) -> None:
        """Open dialog to edit contact information for selected student"""
        selection = contact_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a student first.")
            return
        
        selected_text = contact_listbox.get(selection[0])
        try:
            student_id = int(selected_text.split("ID: ")[1].rstrip(")"))
        except:
            return
        
        student_name = self.db.conn.execute("SELECT name FROM students WHERE id = ?", (student_id,)).fetchone()[0]
        contact = self.db.conn.execute("SELECT * FROM contact_info WHERE student_id = ?", (student_id,)).fetchone()
        
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Contact Info - {student_name}")
        dialog.geometry("500x450")
        
        # Create form
        form_frame = tk.Frame(dialog)
        form_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Labels and entries
        fields = [
            ("Phone:", "phone"),
            ("Email:", "email"),
            ("Address:", "address"),
            ("Guardian Name:", "guardian_name"),
            ("Guardian Phone:", "guardian_phone"),
            ("Notes:", "notes")
        ]
        
        entries = {}
        values = {
            "phone": contact[2] if contact else "",
            "email": contact[3] if contact else "",
            "address": contact[4] if contact else "",
            "guardian_name": contact[5] if contact else "",
            "guardian_phone": contact[6] if contact else "",
            "notes": contact[7] if contact else ""
        }
        
        for label, field in fields:
            tk.Label(form_frame, text=label, font=("Arial", 9)).pack(anchor="w", pady=(10, 0))
            entry = tk.Entry(form_frame, width=50, font=("Arial", 9))
            entry.insert(0, values[field])
            entry.pack(fill="x", pady=(0, 5))
            entries[field] = entry
        
        # Save button
        def save_contact():
            if not contact:
                self.db.conn.execute("""
                    INSERT INTO contact_info (student_id, phone, email, address, guardian_name, guardian_phone, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    student_id,
                    entries["phone"].get(),
                    entries["email"].get(),
                    entries["address"].get(),
                    entries["guardian_name"].get(),
                    entries["guardian_phone"].get(),
                    entries["notes"].get()
                ))
            else:
                self.db.conn.execute("""
                    UPDATE contact_info 
                    SET phone=?, email=?, address=?, guardian_name=?, guardian_phone=?, notes=?
                    WHERE student_id=?
                """, (
                    entries["phone"].get(),
                    entries["email"].get(),
                    entries["address"].get(),
                    entries["guardian_name"].get(),
                    entries["guardian_phone"].get(),
                    entries["notes"].get(),
                    student_id
                ))
            self.db.conn.commit()
            messagebox.showinfo("Success", "Contact information saved successfully.")
            dialog.destroy()
            self._show_contact_info(contact_listbox)
        
        button_frame = tk.Frame(dialog)
        button_frame.pack(fill="x", padx=10, pady=10)
        tk.Button(button_frame, text="Save", width=15, command=save_contact).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancel", width=15, command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def _generate_attendance_report(self) -> None:
        """Generate attendance report for selected date range"""
        try:
            from_date_str = self.report_from_date_var.get().strip()
            to_date_str = self.report_to_date_var.get().strip()
            
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
            
            if from_date > to_date:
                messagebox.showerror("Error", "From date must be before To date")
                return
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD")
            return
        
        # Clear previous data
        for item in self.report_table.get_children():
            self.report_table.delete(item)

        # Get all students
        students = self.db.list_students()
        if not students:
            messagebox.showinfo("Info", "No students found")
            return

        def _parse_class_start(start_str: str, arrival: datetime) -> Optional[datetime]:
            if not start_str:
                return None
            for fmt in ("%I:%M %p", "%H:%M", "%H:%M:%S"):
                try:
                    t = datetime.strptime(start_str, fmt).time()
                    return datetime.combine(arrival.date(), t) + timedelta(minutes=15)
                except Exception:
                    continue
            return None

        # Generate statistics for each student
        for student_id, student_name in students:
            class_ids = {cid for cid, in self.db.conn.execute(
                "SELECT class_id FROM class_students WHERE student_id = ?", (student_id,)
            ).fetchall()}

            # Attendance records in range
            records = self.db.conn.execute(
                "SELECT class_id, timestamp FROM attendance WHERE student_id = ? AND timestamp >= ? AND timestamp <= ?",
                (student_id, f"{from_date_str}T00:00:00", f"{to_date_str}T23:59:59")
            ).fetchall()

            present_count = 0
            late_count = 0

            for class_id, ts_str in records:
                try:
                    arrival_dt = datetime.fromisoformat(ts_str)
                except Exception:
                    try:
                        arrival_dt = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        continue

                start_row = self.db.conn.execute("SELECT start_time FROM classes WHERE id=?", (class_id,)).fetchone()
                start_str = start_row[0] if start_row else ""
                late_cutoff = _parse_class_start(start_str, arrival_dt)

                if late_cutoff and arrival_dt > late_cutoff:
                    late_count += 1
                else:
                    present_count += 1

            total_attended = present_count + late_count
            expected_classes = max(len(class_ids), total_attended)
            absent_count = max(0, expected_classes - total_attended)
            attendance_rate = (present_count / expected_classes * 100) if expected_classes else 0

            if attendance_rate >= 80:
                tag = "good"
            elif attendance_rate >= 60:
                tag = "warning"
            else:
                tag = "bad"

            self.report_table.insert(
                "",
                tk.END,
                values=(student_name, expected_classes, present_count, late_count, absent_count, f"{attendance_rate:.1f}%"),
                tags=(tag,),
            )

        self._log(f"✓ Generated report for {from_date_str} to {to_date_str}")
    
    def _export_report_csv(self) -> None:
        """Export the generated report as CSV"""
        try:
            from_date_str = self.report_from_date_var.get().strip()
            to_date_str = self.report_to_date_var.get().strip()
        except:
            messagebox.showerror("Error", "Generate report first")
            return
        
        # Get all data from table
        rows = []
        for item in self.report_table.get_children():
            values = self.report_table.item(item)["values"]
            rows.append(values)
        
        if not rows:
            messagebox.showinfo("Info", "No data to export. Generate report first.")
            return
        
        default_name = f"attendance_report_{from_date_str}_to_{to_date_str}.csv"
        filename = filedialog.asksaveasfilename(
            title="Save Attendance Report CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not filename:
            return
        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Attendance Report", f"From: {from_date_str} To: {to_date_str}"])
            writer.writerow([])
            writer.writerow(["Student Name", "Total Classes", "Present", "Late", "Absent", "Attendance Rate"])
            writer.writerows(rows)
            
            # Add summary
            writer.writerow([])
            writer.writerow(["SUMMARY"])
            total_students = len(rows)
            avg_attendance = sum(float(row[5].rstrip("%")) for row in rows) / total_students if rows else 0
            writer.writerow([f"Total Students: {total_students}"])
            writer.writerow([f"Average Attendance Rate: {avg_attendance:.1f}%"])
        
        self._log(f"Exported report to {filename}")
        messagebox.showinfo("Success", f"Report saved to {filename}")

    def _generate_student_pdf_report(self) -> None:
        """Generate a PDF report card for a single student"""
        if not HAS_FPDF:
            messagebox.showerror("Error", "fpdf library not installed. Please run: pip install fpdf")
            return
            
        selection = self.pdf_student_var.get()
        if not selection:
            messagebox.showwarning("Warning", "Please select a student first")
            return
            
        try:
            student_id = int(selection.split("#")[1])
            student_name = selection.split("#")[0].strip()
        except:
            messagebox.showerror("Error", "Invalid student selection")
            return
            
        # Get all classes for the student
        classes = self.db.conn.execute(
            "SELECT c.id, c.name FROM classes c JOIN class_students cs ON c.id = cs.class_id WHERE cs.student_id = ?", 
            (student_id,)
        ).fetchall()
        
        if not classes:
            messagebox.showinfo("Info", "Student is not enrolled in any classes")
            return
            
        # Get all warnings for this student
        all_warnings = self.db.get_all_warnings_summary()
        student_warnings = [w for w in all_warnings if w[0] == student_id]
        
        # Select save location
        default_name = f"Report_Card_{student_name.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
        filename = filedialog.asksaveasfilename(
            title="Save PDF Report Card",
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if not filename:
            return
            
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Title
            pdf.set_font("Arial", "B", 20)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(0, 15, "Student Attendance Report Card", ln=True, align="C")
            
            pdf.set_font("Arial", "I", 12)
            pdf.set_text_color(127, 140, 141)
            pdf.cell(0, 10, f"Generated on: {datetime.now().strftime('%Y-%m-%d')}", ln=True, align="C")
            
            pdf.ln(10)
            
            # Student Info
            pdf.set_font("Arial", "B", 14)
            pdf.set_text_color(44, 62, 80)
            pdf.cell(40, 10, "Student Name:", 0, 0)
            pdf.set_font("Arial", "", 14)
            pdf.cell(0, 10, student_name, 0, 1)
            
            pdf.set_font("Arial", "B", 14)
            pdf.cell(40, 10, "Student ID:", 0, 0)
            pdf.set_font("Arial", "", 14)
            pdf.cell(0, 10, str(student_id), 0, 1)
            
            pdf.ln(10)
            
            # Table Header
            pdf.set_fill_color(52, 152, 219)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", "B", 12)
            pdf.cell(60, 10, "Subject / Class", 1, 0, "C", fill=True)
            pdf.cell(30, 10, "Total", 1, 0, "C", fill=True)
            pdf.cell(30, 10, "Attended", 1, 0, "C", fill=True)
            pdf.cell(30, 10, "Absent", 1, 0, "C", fill=True)
            pdf.cell(40, 10, "Warning Level", 1, 1, "C", fill=True)
            
            # Table Data
            pdf.set_font("Arial", "", 11)
            
            fill = False
            for w in student_warnings:
                _, _, c_name, total, attended, absent, warning = w
                
                pdf.set_text_color(0, 0, 0)
                if fill:
                    pdf.set_fill_color(236, 240, 241)
                else:
                    pdf.set_fill_color(255, 255, 255)
                
                pdf.cell(60, 10, c_name[:25], 1, 0, "L", fill=True)
                pdf.cell(30, 10, str(total), 1, 0, "C", fill=True)
                
                # Attended
                pdf.set_text_color(39, 174, 96)
                pdf.cell(30, 10, str(attended), 1, 0, "C", fill=True)
                
                # Absent
                if absent > 0:
                    pdf.set_text_color(192, 57, 43)
                else:
                    pdf.set_text_color(0, 0, 0)
                pdf.cell(30, 10, str(absent), 1, 0, "C", fill=True)
                
                # Warning
                if warning:
                    pdf.set_text_color(192, 57, 43)
                    pdf.set_font("Arial", "B", 11)
                else:
                    pdf.set_text_color(39, 174, 96)
                    pdf.set_font("Arial", "", 11)
                    warning = "OK"
                
                pdf.cell(40, 10, warning, 1, 1, "C", fill=True)
                
                pdf.set_font("Arial", "", 11)
                fill = not fill
                
            pdf.output(filename)
            self._log(f"Exported PDF report to {filename}")
            messagebox.showinfo("Success", f"PDF Report Card saved to:\n{filename}")
            
        except Exception as exc:
            messagebox.showerror("Export Error", f"Failed to generate PDF:\n{exc}")

    def _backup_database(self) -> None:
        """Create a backup copy of the database"""
        try:
            backup_name = f"attendance_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            target = filedialog.asksaveasfilename(title="Save Backup As", defaultextension=".db",
                                                 initialfile=backup_name,
                                                 filetypes=[("SQLite DB", "*.db"), ("All Files", "*.*")])
            if not target:
                return
            self.db.conn.commit()
            shutil.copyfile(DB_PATH, target)
            self._log(f"✓ Backup saved to {target}")
            messagebox.showinfo("Backup", "Backup created successfully!")
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Backup", f"Could not create backup:\n{exc}")

    def _restore_database(self) -> None:
        """Restore database from a backup file"""
        try:
            source = filedialog.askopenfilename(title="Select Backup File",
                                                filetypes=[("SQLite DB", "*.db"), ("All Files", "*.*")])
            if not source:
                return
            if not messagebox.askyesno("Restore", "This will replace the current data. Continue?"):
                return
            self.db.conn.commit()
            self.db.conn.close()
            shutil.copyfile(source, DB_PATH)
            self.db = Database(DB_PATH)
            self._refresh_student_list()
            self._refresh_teacher_list()
            self._refresh_class_list()
            if hasattr(self, "contact_listbox_ref"):
                self._refresh_contact_list(self.contact_listbox_ref)
            self._refresh_class_dropdown()
            self._log(f"✓ Database restored from {source}")
            messagebox.showinfo("Restore", "Database restored successfully.")
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Restore", f"Restore failed:\n{exc}")

    def _load_selected_date(self) -> None:
        """Load history for the date selected from dropdown"""
        selected = self.available_dates_var.get()
        if selected and selected != "No dates available":
            self.history_date_var.set(selected)
            self._load_attendance_history()
    
    def _refresh_available_dates(self) -> None:
        """Refresh the dropdown with all available attendance dates"""
        dates = self.db.get_all_attendance_dates()
        
        menu = self.dates_dropdown["menu"]
        menu.delete(0, tk.END)
        
        if dates:
            for date_str in dates:
                menu.add_command(label=date_str, 
                               command=tk._setit(self.available_dates_var, date_str))
            self.available_dates_var.set(dates[0])  # Set to most recent
            self._log(f"✓ Found {len(dates)} date(s) with attendance records")
        else:
            menu.add_command(label="No dates available", 
                           command=tk._setit(self.available_dates_var, "No dates available"))
            self.available_dates_var.set("No dates available")
            self._log("📭 No attendance records found")

    def _load_attendance_history(self) -> None:
        """Load attendance history for the selected date"""
        date_str = self.history_date_var.get().strip()
        
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD")
            return
        
        # Clear previous data
        for item in self.history_table.get_children():
            self.history_table.delete(item)
        
        # Get all classes that had sessions on this date
        date_start = f"{date_str}T00:00:00"
        date_end = f"{date_str}T23:59:59"
        
        # Get classes with attendance records on this date
        classes_with_attendance = self.db.conn.execute("""
            SELECT DISTINCT c.id, c.name, c.start_time, COALESCE(a.session_number, 0) AS session_number
            FROM classes c
            JOIN attendance a ON c.id = a.class_id
            WHERE a.timestamp >= ? AND a.timestamp <= ?
            ORDER BY c.name, session_number
        """, (date_start, date_end)).fetchall()
        
        if not classes_with_attendance:
            self.history_stats_var.set(f"📭 No attendance records found for {date_str}")
            self._log(f"No attendance records found for {date_str}")
            return
        
        total_classes = len(classes_with_attendance)
        total_students_present = 0
        total_students_late = 0
        total_students_absent = 0

        # Process each class/session combination that has attendance
        for class_id, class_name, start_time_str, session_num in classes_with_attendance:
            class_students = self.db.list_class_students(class_id)
            student_dict = {sid: name for sid, name in class_students}

            # Display class name with session number when available
            display_name = f"{class_name} (Session #{session_num})" if session_num else class_name

            attendance_records = self.db.conn.execute("""
                SELECT s.id, s.name, a.timestamp
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE a.class_id = ? AND a.timestamp >= ? AND a.timestamp <= ?
                  AND (? = 0 OR a.session_number = ?)
                ORDER BY a.timestamp
            """, (class_id, date_start, date_end, session_num, session_num)).fetchall()

            # Teacher arrivals for the class on this date
            teacher_records = self.db.conn.execute("""
                SELECT t.name, ta.timestamp
                FROM teacher_attendance ta
                JOIN teachers t ON ta.teacher_id = t.id
                WHERE ta.class_id = ? AND ta.timestamp >= ? AND ta.timestamp <= ?
                ORDER BY ta.timestamp
            """, (class_id, date_start, date_end)).fetchall()
            
            teacher_name = teacher_records[0][0] if teacher_records else "No Teacher"
            teacher_time = datetime.fromisoformat(teacher_records[0][1]) if teacher_records else None
            
            # Calculate late cutoff (15 minutes after scheduled start or 5 min after teacher arrival)
            late_cutoff = None
            if teacher_time:
                class_actual_start = teacher_time + timedelta(minutes=5)
                late_cutoff = class_actual_start
            elif start_time_str:
                try:
                    # Parse start time
                    for fmt in ("%I:%M %p", "%H:%M", "%H:%M:%S"):
                        try:
                            t = datetime.strptime(start_time_str, fmt).time()
                            late_cutoff = datetime.combine(selected_date.date(), t) + timedelta(minutes=15)
                            break
                        except:
                            continue
                except:
                    pass
            
            # Track who attended
            attended_ids = set()
            
            # Add class header row with session number
            display_name = f"{class_name} (Session #{session_num})" if session_num > 0 else class_name
            self.history_table.insert("", tk.END, 
                                     values=(f"━━━ {display_name} ━━━", "", "", "", teacher_name),
                                     tags=("header_row",))
            
            # Process each attendance record
            for student_id, student_name, timestamp_str in attendance_records:
                attended_ids.add(student_id)
                arrival_time = datetime.fromisoformat(timestamp_str)
                
                # Determine status
                if late_cutoff and arrival_time > late_cutoff:
                    status = "⏰ Late"
                    tag = "late"
                    total_students_late += 1
                else:
                    status = "✓ Present"
                    tag = "present"
                    total_students_present += 1
                
                self.history_table.insert("", tk.END,
                                         values=("  " + display_name, 
                                                student_name, 
                                                status, 
                                                arrival_time.strftime("%H:%M:%S"),
                                                ""),
                                         tags=(tag,))
            
            # Add absent students
            for student_id, student_name in class_students:
                if student_id not in attended_ids:
                    self.history_table.insert("", tk.END,
                                             values=("  " + display_name,
                                                    student_name,
                                                    "❌ Absent",
                                                    "-",
                                                    ""),
                                             tags=("absent",))
                    total_students_absent += 1
            
            # Add unknown visitors for this session
            unknown_visitors = self.db.conn.execute("""
                SELECT id, visit_count, timestamp
                FROM unknown_visitors
                WHERE class_id = ? AND timestamp >= ? AND timestamp <= ?
                ORDER BY timestamp
            """, (class_id, date_start, date_end)).fetchall()
            
            for visitor_id, visit_count, ts_str in unknown_visitors:
                ts = datetime.fromisoformat(ts_str)
                self.history_table.insert("", tk.END,
                                         values=("  " + display_name,
                                                f"Unknown Visitor #{visitor_id}",
                                                f"👤 Visitor (×{visit_count})",
                                                ts.strftime("%H:%M:%S"),
                                                ""),
                                         tags=("late",))
            
            # Add spacing
            self.history_table.insert("", tk.END, values=("", "", "", "", ""))
        
        # Update statistics
        total_records = total_students_present + total_students_late + total_students_absent
        stats_text = (f"📊 Summary for {date_str}: "
                     f"{total_classes} class(es) | "
                     f"✓ {total_students_present} present | "
                     f"⏰ {total_students_late} late | "
                     f"❌ {total_students_absent} absent | "
                     f"Total: {total_records} records")
        
        self.history_stats_var.set(stats_text)
        self._log(f"✓ Loaded history for {date_str}: {total_classes} class(es), {total_records} records")
    
    def _export_history_csv(self) -> None:
        """Export attendance history to CSV"""
        date_str = self.history_date_var.get().strip()
        
        # Check if there's data to export
        if not self.history_table.get_children():
            messagebox.showinfo("Export", "No data to export. Please load attendance history first.")
            return
        
        # Collect data from treeview
        rows = []
        for item in self.history_table.get_children():
            values = self.history_table.item(item)["values"]
            if values and values[0]:  # Skip empty rows
                rows.append(values)
        
        if not rows:
            messagebox.showinfo("Export", "No data to export.")
            return
        
        # Create filename with file dialog
        default_name = f"attendance_history_{date_str}.csv"
        filename = filedialog.asksaveasfilename(
            title="Save History CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not filename:
            return
        
        try:
            with open(filename, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([f"Attendance History for {date_str}"])
                writer.writerow([])
                writer.writerow(["Class Name", "Student Name", "Status", "Time", "Teacher"])
                writer.writerows(rows)
                
                # Add summary
                writer.writerow([])
                writer.writerow(["Summary"])
                writer.writerow([self.history_stats_var.get()])
            
            self._log(f"✓ Exported history to {filename}")
            messagebox.showinfo("Export Success", f"History exported to:\n{filename}")
        except Exception as exc:
            messagebox.showerror("Export Error", f"Failed to export:\n{exc}")
    
    def _export_history_excel(self) -> None:
        """Export attendance history to Excel with formatting"""
        if not HAS_OPENPYXL:
            messagebox.showerror("Excel", "openpyxl not installed. Run: pip install openpyxl")
            return
        
        date_str = self.history_date_var.get().strip()
        
        # Check if there's data to export
        if not self.history_table.get_children():
            messagebox.showinfo("Export", "No data to export. Please load attendance history first.")
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"History {date_str}"
            
            # Add title
            ws.append([f"Attendance History for {date_str}"])
            ws.merge_cells("A1:E1")
            title_cell = ws["A1"]
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.append([])
            
            # Add headers
            ws.append(["Class Name", "Student Name", "Status", "Time", "Teacher"])
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Add data with colors
            for item in self.history_table.get_children():
                values = self.history_table.item(item)["values"]
                tags = self.history_table.item(item)["tags"]
                
                if values and values[0]:  # Skip empty rows
                    ws.append(values)
                    row_num = ws.max_row
                    
                    # Color code based on status
                    if "present" in tags:
                        for col in ["A", "B", "C", "D", "E"]:
                            ws[f"{col}{row_num}"].fill = PatternFill(
                                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                            )
                            ws[f"{col}{row_num}"].font = Font(color="006100")
                    elif "late" in tags:
                        for col in ["A", "B", "C", "D", "E"]:
                            ws[f"{col}{row_num}"].fill = PatternFill(
                                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                            )
                            ws[f"{col}{row_num}"].font = Font(color="9C6500")
                    elif "absent" in tags:
                        for col in ["A", "B", "C", "D", "E"]:
                            ws[f"{col}{row_num}"].fill = PatternFill(
                                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                            )
                            ws[f"{col}{row_num}"].font = Font(color="9C0006")
                    elif "header_row" in tags:
                        for col in ["A", "B", "C", "D", "E"]:
                            ws[f"{col}{row_num}"].fill = PatternFill(
                                start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"
                            )
                            ws[f"{col}{row_num}"].font = Font(bold=True, color="FFFFFF")
            
            # Add summary
            ws.append([])
            ws.append(["Summary:"])
            ws.append([self.history_stats_var.get()])
            
            # Set column widths
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 20
            
            # Save file with file dialog
            default_name = f"attendance_history_{date_str}.xlsx"
            filename = filedialog.asksaveasfilename(
                title="Save History Excel",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
            )
            if not filename:
                return
            wb.save(filename)
            
            self._log(f"✓ Exported history to {filename}")
            messagebox.showinfo("Export Success", f"History exported to:\n{filename}\n\nWith color coding:\n✓ Green = Present\n⏰ Yellow = Late\n❌ Red = Absent")
        except Exception as exc:
            messagebox.showerror("Export Error", f"Failed to export:\n{exc}")

    def _refresh_warnings_dropdown(self) -> None:
        """Refresh the class filter dropdown in warnings tab"""
        classes = self.db.list_classes()
        menu = self.warning_class_dropdown["menu"]
        menu.delete(0, tk.END)

        menu.add_command(label="All Classes",
                        command=tk._setit(self.warning_class_filter_var, "All Classes"))

        for class_id, class_name in classes:
            label = f"{class_name} #{class_id}"
            menu.add_command(label=label,
                           command=tk._setit(self.warning_class_filter_var, label))

        self.warning_class_filter_var.set("All Classes")

    def _refresh_warnings_tab(self) -> None:
        """Refresh the warnings table with per-class warning data"""
        for item in self.warnings_table.get_children():
            self.warnings_table.delete(item)

        filter_value = self.warning_class_filter_var.get()

        if filter_value == "All Classes":
            all_data = self.db.get_all_warnings_summary()
        else:
            try:
                class_id = int(filter_value.split("#")[1])
                summary = self.db.get_warning_summary_for_class(class_id)
                class_name = filter_value.split("#")[0].strip()
                all_data = [(sid, sname, class_name, total, attended, absent, warning)
                            for sid, sname, total, attended, absent, warning in summary]
            except Exception:
                all_data = self.db.get_all_warnings_summary()

        warning_count = 0
        for student_id, student_name, class_name, total_lect, attended, absent, warning in all_data:
            if warning:
                if "Last" in warning or "last" in warning:
                    tag = "warning_last"
                    display_warning = "🚫 Last Warning"
                elif "2nd" in warning:
                    tag = "warning_2nd"
                    display_warning = "⚠️⚠️ 2nd Warning"
                elif "1st" in warning:
                    tag = "warning_1st"
                    display_warning = "⚠️ 1st Warning"
                else:
                    tag = "warning_1st"
                    display_warning = warning
                warning_count += 1
            else:
                tag = "no_warning"
                display_warning = "✅ OK"

            self.warnings_table.insert("", tk.END,
                                       values=(student_name, class_name, total_lect, attended, absent, display_warning),
                                       tags=(tag,))

        total_rows = len(all_data)
        self.warnings_count_label.config(
            text=f"📋 {total_rows} record(s) | ⚠️ {warning_count} warning(s)"
        )

        self._log(f"✓ Warnings refreshed: {total_rows} records, {warning_count} warnings")

    def _export_warnings_excel(self) -> None:
        """Export warnings to Excel with file dialog"""
        if not HAS_OPENPYXL:
            messagebox.showerror("Excel", "openpyxl not installed. Run: pip install openpyxl")
            return

        all_data = self.db.get_all_warnings_summary()

        if not all_data:
            messagebox.showinfo("Export", "No warning data to export.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Warnings"

        # Title row
        ws.append(["Student Warnings & Absence Tracking"])
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.append(["Warning Rules: 2 Absences = 1st Warning | 3 Absences = 2nd Warning | 4+ Absences = Last Warning"])
        ws.merge_cells("A2:F2")
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.append([])

        headers = ["Student Name", "Class Name", "Total Lectures", "Attended", "Absent", "Warning Level"]
        ws.append(headers)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for student_id, student_name, class_name, total_lect, attended, absent, warning in all_data:
            ws.append([student_name, class_name, total_lect, attended, absent, warning or "OK"])
            row_num = ws.max_row

            if warning:
                if "Last" in warning or "last" in warning:
                    color, font_color = "FF0000", "FFFFFF"
                elif "2nd" in warning:
                    color, font_color = "FF6600", "FFFFFF"
                else:
                    color, font_color = "FFC107", "000000"
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col}{row_num}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    ws[f"{col}{row_num}"].font = Font(color=font_color, bold=True)
            else:
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col}{row_num}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f"{col}{row_num}"].font = Font(color="006100")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 18

        # Per-class sheets
        classes = self.db.list_classes()
        for class_id, class_name in classes:
            safe_name = class_name[:28].replace("/", "-").replace("\\", "-")
            ws_class = wb.create_sheet(safe_name)

            ws_class.append([f"Warnings for: {class_name}"])
            ws_class.merge_cells("A1:E1")
            ws_class["A1"].font = Font(size=14, bold=True, color="FFFFFF")
            ws_class["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws_class["A1"].alignment = Alignment(horizontal="center")

            ws_class.append([])
            ws_class.append(["Student Name", "Total Lectures", "Attended", "Absent", "Warning"])
            for cell in ws_class[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            class_summary = self.db.get_warning_summary_for_class(class_id)
            for _sid, student_name, total_lect, attended, absent, warning in class_summary:
                ws_class.append([student_name, total_lect, attended, absent, warning or "OK"])
                row_num = ws_class.max_row
                if warning:
                    if "Last" in warning or "last" in warning:
                        color, font_color = "FF0000", "FFFFFF"
                    elif "2nd" in warning:
                        color, font_color = "FF6600", "FFFFFF"
                    else:
                        color, font_color = "FFC107", "000000"
                    for col in ["A", "B", "C", "D", "E"]:
                        ws_class[f"{col}{row_num}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        ws_class[f"{col}{row_num}"].font = Font(color=font_color, bold=True)

            ws_class.column_dimensions["A"].width = 25
            ws_class.column_dimensions["B"].width = 15
            ws_class.column_dimensions["C"].width = 12
            ws_class.column_dimensions["D"].width = 12
            ws_class.column_dimensions["E"].width = 18

        default_name = f"student_warnings_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Save Warnings Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if not filename:
            return

        wb.save(filename)
        self._log(f"✓ Exported warnings to {filename}")
        messagebox.showinfo("Export Success", f"Warnings exported to:\n{filename}\n\nIncludes:\n- All warnings summary\n- Per-class warning sheets")

    def _refresh_dashboard(self) -> None:
        """Refresh the charts in the dashboard tab"""
        if not HAS_MATPLOTLIB:
            messagebox.showerror("Error", "matplotlib is not installed. Please run: pip install matplotlib")
            return
            
        # Clear existing charts
        for widget in self.charts_frame.winfo_children():
            widget.destroy()
            
        try:
            # Gather data
            all_warnings = self.db.get_all_warnings_summary()
            
            total_students = len(self.db.list_students())
            if total_students == 0:
                tk.Label(self.charts_frame, text="No students to display.", font=("Segoe UI", 14), bg="white").pack(expand=True)
                return
                
            total_attended = sum(w[4] for w in all_warnings)
            total_absent = sum(w[5] for w in all_warnings)
            
            # Pie Chart Data
            labels_pie = []
            sizes_pie = []
            colors_pie = []
            
            if total_attended > 0:
                labels_pie.append('Present')
                sizes_pie.append(total_attended)
                colors_pie.append('#2ecc71')
                
            if total_absent > 0:
                labels_pie.append('Absent')
                sizes_pie.append(total_absent)
                colors_pie.append('#e74c3c')
                
            if not sizes_pie:
                sizes_pie = [1]
                labels_pie = ['No Data']
                colors_pie = ['#bdc3c7']
                
            # Bar Chart Data (Absences per class)
            class_absences = {}
            for w in all_warnings:
                c_name = w[2]
                absent = w[5]
                class_absences[c_name] = class_absences.get(c_name, 0) + absent
                
            # Sort for bar chart
            sorted_classes = sorted(class_absences.items(), key=lambda x: x[1], reverse=True)[:5] # Top 5
            bar_labels = [k[:10]+".." if len(k)>12 else k for k, v in sorted_classes]
            bar_values = [v for k, v in sorted_classes]
            
            # Create Figure
            fig = Figure(figsize=(10, 5), dpi=100)
            fig.patch.set_facecolor('#ffffff')
            
            # Plot 1: Pie Chart (Overall Attendance)
            ax1 = fig.add_subplot(121)
            ax1.pie(sizes_pie, labels=labels_pie, colors=colors_pie, autopct='%1.1f%%', startangle=90)
            ax1.set_title('Overall Attendance vs Absence')
            
            # Plot 2: Bar Chart (Most Absences by Class)
            ax2 = fig.add_subplot(122)
            if bar_values:
                bars = ax2.bar(bar_labels, bar_values, color='#3498db')
                ax2.set_title('Top 5 Classes by Absences')
                ax2.set_ylabel('Total Absences')
                ax2.tick_params(axis='x', rotation=45)
            else:
                ax2.text(0.5, 0.5, 'No absence data', ha='center', va='center')
                ax2.set_title('Classes by Absences')
                
            fig.tight_layout()
            
            # Embed in Tkinter
            canvas = FigureCanvasTkAgg(fig, master=self.charts_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            self._log("✓ Dashboard charts refreshed")
        except Exception as e:
            tk.Label(self.charts_frame, text=f"Error rendering charts: {str(e)}", fg="red", bg="white").pack(expand=True)
            self._log(f"Error rendering dashboard: {e}")

    def _apply_settings(self) -> None:
        """Persist tolerance, camera index, and theme settings"""
        global ENCODING_TOLERANCE
        try:
            ENCODING_TOLERANCE = float(self.encoding_tolerance_var.get())
        except Exception:
            messagebox.showerror("Settings", "Invalid tolerance value")
            return
        # Camera index already bound to self.camera_index_var
        self._apply_theme()
        self._log(f"✓ Settings applied (tolerance={ENCODING_TOLERANCE:.2f}, camera={self.camera_index_var.get()}, theme={self.theme_var.get()})")
        messagebox.showinfo("Settings", "Settings saved")

    def _apply_theme(self) -> None:
        """Apply a cleaner theme without fighting native widget colors."""
        theme_name = getattr(self, "theme_var", tk.StringVar(value="light")).get()

        # Palettes keep contrast while giving a subtle frame tint
        palettes = {
            "dark": {"bg": "#0f172a", "card": "#111827", "input": "#192233", "fg": "#e5e7eb", "accent": "#22c55e"},
            "blue": {"bg": "#0b1220", "card": "#10192f", "input": "#15203a", "fg": "#e2e8f0", "accent": "#38bdf8"},
            "light": {"bg": "#f5f7fb", "card": "#ffffff", "input": "#fdfefe", "fg": "#1f2937", "accent": "#2563eb"},
        }
        palette = palettes.get(theme_name, palettes["light"])

        try:
            self.root.configure(bg=palette["bg"])
            self._style_ttk(palette)
            self._theme_widgets(self.root, palette)
        except Exception:
            pass

    def _theme_widgets(self, widget, palette: Dict[str, str]) -> None:
        """Recursively apply background/foreground to common widgets without breaking ttk defaults."""
        try:
            import tkinter.ttk as ttk  # local import to avoid top-level dependency
        except Exception:
            ttk = None

        # Apply to this widget
        wtype = widget.winfo_class()
        try:
            if wtype in {"Frame", "Labelframe"}:
                widget.configure(bg=palette["card"])
            elif wtype in {"Label"}:
                widget.configure(bg=palette["card"], fg=palette["fg"])
            elif wtype in {"Button"}:
                widget.configure(bg=palette["accent"], fg="#ffffff", activebackground=palette["accent"], activeforeground="#ffffff", relief=tk.FLAT, bd=0, padx=8, pady=6)
            elif wtype in {"Menubutton"}:
                widget.configure(bg=palette["card"], fg=palette["fg"], activebackground=palette["accent"], activeforeground="#ffffff", relief=tk.FLAT, bd=0, padx=8, pady=6)
                try:
                    menu = widget["menu"]
                    menu.configure(bg=palette["card"], fg=palette["fg"], activebackground=palette["accent"], activeforeground="#ffffff", relief=tk.FLAT, bd=0)
                except Exception:
                    pass
            elif wtype in {"Checkbutton", "Radiobutton"}:
                widget.configure(bg=palette["card"], fg=palette["fg"], activebackground=palette["card"], selectcolor=palette["accent"], highlightthickness=0)
            elif wtype in {"Entry", "Listbox", "Text"}:
                input_bg = palette.get("input", palette["card"])
                widget.configure(
                    bg=input_bg,
                    fg=palette["fg"],
                    insertbackground=palette["fg"],
                    highlightbackground=palette["card"],
                    highlightcolor=palette["accent"],
                    relief=tk.FLAT,
                    bd=1,
                )
                if wtype == "Listbox":
                    widget.configure(selectbackground=palette["accent"], selectforeground="#ffffff")
            elif wtype in {"TFrame", "TLabelframe"} and ttk:
                style = ttk.Style()
                style.configure("Themed.TFrame", background=palette["card"])
                widget.configure(style="Themed.TFrame")
            elif wtype in {"Treeview"} and ttk:
                widget.configure(style="Themed.Treeview")
        except Exception:
            pass

        # Recurse
        for child in widget.winfo_children():
            self._theme_widgets(child, palette)

    def _style_ttk(self, palette: Dict[str, str]) -> None:
        """Configure ttk styles (Notebook, Treeview, Buttons) to match palette."""
        try:
            import tkinter.ttk as ttk
        except Exception:
            return

        style = ttk.Style()
        current_theme = style.theme_use()
        style.theme_use(current_theme)

        # Notebook
        style.configure("Themed.TNotebook", background=palette["bg"], borderwidth=0)
        style.configure("Themed.TNotebook.Tab", padding=(14, 8), background=palette["card"], foreground=palette["fg"], borderwidth=0)
        style.map("Themed.TNotebook.Tab",
                  background=[("selected", palette["accent"]), ("!selected", palette["card"])],
                  foreground=[("selected", "#ffffff"), ("!selected", palette["fg"])])

        # Label style for cleaner headings
        style.configure("Themed.TLabel", background=palette["card"], foreground=palette["fg"], padding=2)

        # Treeview
        style.configure("Themed.Treeview", background=palette["card"], fieldbackground=palette["card"], foreground=palette["fg"], bordercolor=palette["card"], rowheight=22)
        style.map("Themed.Treeview", background=[("selected", palette["accent"])], foreground=[("selected", "#ffffff")])

        # Buttons
        style.configure("Themed.TButton", background=palette["accent"], foreground="#ffffff", padding=(10, 6), borderwidth=0, focusthickness=3, focuscolor=palette["accent"])
        style.map("Themed.TButton",
                  background=[("active", palette["accent"]), ("disabled", palette["card"])],
                  foreground=[("disabled", "#999999")])

        # Entry / Combobox
        input_bg = palette.get("input", palette["card"])
        style.configure("Themed.TEntry", fieldbackground=input_bg, foreground=palette["fg"], bordercolor=palette["card"], padding=6)
        style.map("Themed.TEntry", fieldbackground=[("active", input_bg)])

        style.configure("Themed.TCombobox", fieldbackground=input_bg, foreground=palette["fg"], bordercolor=palette["card"], padding=6)
        style.map("Themed.TCombobox",
              fieldbackground=[("readonly", input_bg)],
              background=[("active", input_bg)],
              foreground=[("readonly", palette["fg"])])

        # Attach styles to notebook widgets
        def _attach_notebook_styles(widget):
            if isinstance(widget, ttk.Notebook):
                widget.configure(style="Themed.TNotebook")
            for child in widget.winfo_children():
                _attach_notebook_styles(child)
        _attach_notebook_styles(self.root)

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    app = FaceAttendanceApp()
    app.run()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
